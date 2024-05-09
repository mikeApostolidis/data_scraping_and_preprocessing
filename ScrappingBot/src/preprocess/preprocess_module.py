import os
import re
import warnings

import numpy as np
import openpyxl
import pandas as pd
from datetime import datetime, date
import warnings
import random


def test_remove_all_files(path, zip_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(path):
        if filename.endswith(".csv") or filename.endswith(".xlsx") or filename.endswith(".rar"):
            # Construct the full path of the file
            file_path = os.path.join(path, filename)
            try:
                # Remove the file
                os.remove(file_path)
            except Exception as e:
                print(
                    f"An error occurred during deleting all csv files on method  test_remove_all_files : {e},{filename}")
    for filename in os.listdir(zip_path):
        if filename.endswith(".rar") or filename.endswith(".xlsx"):
            # Construct the full path of the file
            file_path = os.path.join(zip_path, filename)
            try:
                # Remove the file
                os.remove(file_path)
            except Exception as e:
                print(
                    f"An error occurred during deleting all csv files on method  test_remove_all_files : {e},{filename}")


def test_remove_empty_spaces_before_after_commas(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Apply the strip method to each element in the DataFrame
                df = df.map(lambda x: x.strip() if isinstance(x, str) else x)
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method  test_remove_empty_spaces_before_after_commas : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def check_if_column_exists(df, column_name, default_value):
    if column_name not in df.columns:
        df[column_name] = default_value


def check_if_value_from_row_exists(df, column_name, row_index, default_value=None):
    if column_name not in df.columns:
        df[column_name] = np.nan  # or df[column_name] = default_value if you prefer a specific default
    if pd.isnull(df.at[row_index, column_name]):
        df.at[row_index, column_name] = default_value


# methods to exclude hidden rows and sheets from excel files, if you dont need them, comment them
# def read_excel_exclude_hidden(filename):
#     # Load workbook
#     workbook = openpyxl.load_workbook(filename, data_only=True)
#
#     data = {}
#
#
#     # Loop through each sheet
#     for sheetname in workbook.sheetnames:
#         sheet = workbook[sheetname]
#
#         # Skip hidden sheets
#         if sheet.sheet_state == 'hidden':
#             print(f"Skipping hidden sheet: {sheetname}, {filename}")
#             continue
#
#         rows = []
#
#         # Loop through each row
#         for row in sheet.iter_rows():
#             # Skip hidden rows
#             if sheet.row_dimensions[row[0].row].hidden:
#                 print(f"Skipping hidden row: {row[0].row} in sheet {sheetname}, {filename}")
#                 continue
#
#             # Append row values to rows
#             rows.append([cell.value for cell in row])
#
#         # Convert rows to DataFrame and store in data
#         if len(rows) > 1:
#             data[sheetname] = pd.DataFrame(rows[1:], columns=rows[0])
#
#     return data
#
# def test_convert_excel_to_csv(folder_path):
#     # Normalize the sheet names to lowercase English
#     exclude_sheet_substrings = ["ΣΤΑΤΙΣΤΙΚΑ".lower(), "ΣΤΑΤ".lower()]
#
#     # Loop through files in the input folder_path
#     for filename in os.listdir(folder_path):
#         if filename.endswith(".xlsx"):
#             # Construct the full path of the input Excel file
#             input_path = os.path.join(folder_path, filename)
#
#             # Read all sheets from the Excel file into a dictionary of DataFrames
#             excel_sheets = read_excel_exclude_hidden(input_path)
#
#             # Remove the excluded sheets
#             for sheet_name in list(excel_sheets.keys()):  # We need to create a copy of the keys because we are modifying the dictionary
#                 if any(substring in sheet_name.lower() for substring in exclude_sheet_substrings):
#                     del excel_sheets[sheet_name]
#
#             # Loop through the remaining sheets and save each as a separate CSV file
#             for sheet_name, df in excel_sheets.items():
#                 # Construct the full path of the output CSV file
#                 output_filename = os.path.splitext(filename)[0] + f"_{sheet_name}.csv"
#                 output_path = os.path.join(folder_path, output_filename)
#
#                 # Save DataFrame to CSV only if it is not empty
#                 if not df.empty:
#                     df.to_csv(output_path, index=False, encoding='utf-8')
#                     # df.to_csv(output_path, index=False, encoding='utf-8')
#             # Delete the original Excel file
#             # os.remove(input_path)

def test_convert_excel_to_csv(folder_path):
    # Normalize the sheet names to lowercase English
    exclude_sheet_substrings = ["ΣΤΑΤΙΣΤΙΚΑ".lower(), "ΣΤΑΤ".lower()]

    # Loop through files in the input folder_path
    for filename in os.listdir(folder_path):
        if filename.endswith(".xlsx"):
            # Construct the full path of the input Excel file
            input_path = os.path.join(folder_path, filename)

            # Read all sheets from the Excel file into a dictionary of DataFrames
            excel_sheets = pd.read_excel(input_path, sheet_name=None)

            # Remove the excluded sheets
            for sheet_name in list(
                    excel_sheets.keys()):  # We need to create a copy of the keys because we're modifying the dictionary
                if any(substring in sheet_name.lower() for substring in exclude_sheet_substrings):
                    del excel_sheets[sheet_name]

            # Loop through the remaining sheets and save each as a separate CSV file
            for sheet_name, df in excel_sheets.items():
                # Construct the full path of the output CSV file
                output_filename = os.path.splitext(filename)[0] + f"_{sheet_name}.csv"
                output_path = os.path.join(folder_path, output_filename)

                # Save DataFrame to CSV only if it is not empty
                if not df.empty:
                    df.to_csv(output_path, index=False, encoding='utf-8')

            # Delete the original Excel file
            os.remove(input_path)


def test_delete_AA(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('Α/Α', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method  test_delete_AA : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def test_delete_AA_ROHS(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('Α/Α ΡΟΗΣ', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method  test_delete_AA_ROHS : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def test_normalize_type(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            # Check and add 'ΤΥΠΟΣ' column if it doesn't exist
            check_if_column_exists(df, 'ΤΥΠΟΣ', '')

            try:
                df['ΤΥΠΟΣ'] = df['ΤΥΠΟΣ ΤΟΠΟΘΕΤΗΣΗΣ']
                df = df.drop(['ΤΥΠΟΣ ΤΟΠΟΘΕΤΗΣΗΣ'], axis=1)
            except Exception as e:

                print(f"An error occurred during preprocessing on method test_normalize_type : {e},{filename}")

            # Check if the column 'ΤΥΠΟΣ' is empty
            if all(df['ΤΥΠΟΣ'] == ''):
                # Check the filename for the keywords
                keywords = ['ΓΕΝΙΚΗΣ', 'ΓΕΝΙΚΗ', 'γενικής', 'γενική', 'γενικη', 'γενικης']
                for keyword in keywords:
                    if keyword in filename:
                        # Assign the corresponding value to the entire 'ΤΥΠΟΣ' column
                        df['ΤΥΠΟΣ'] = 'ΓΕΝΙΚΗΣ ΠΑΙΔΕΙΑΣ'
                        break  # Exit the loop after finding the first matching keyword

            type_mapping = {
                'ΓΕΝΙΚΗΣ ΠΑΙΔΕΙΑΣ': 'ΓΕΝΙΚΗΣ',
                'ΜΟΥΣΙΚΟ ΣΧΟΛΕΙΟ': 'ΜΟΥΣΙΚΟ',
            }

            try:

                # Use the mapping dictionary to replace values in the ΤΥΠΟΣ column
                df['ΤΥΠΟΣ'] = df['ΤΥΠΟΣ'].replace(type_mapping)

                # For all other values, set the value to 'ΕΙΔΙΚΗΣ'
                df['ΤΥΠΟΣ'] = df['ΤΥΠΟΣ'].apply(lambda x: x if x in ['ΓΕΝΙΚΗΣ', 'ΜΟΥΣΙΚΟ'] else 'ΕΙΔΙΚΗΣ')
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_normalize_type : {e},{filename}")

            # Iterate over each row in the DataFrame
            for index, row in df.iterrows():
                try:
                    # Check if the value 'Μουσικό Σχολείο' exists in the column 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'
                    if 'Μουσικό Σχολείο' in str(row['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ']):
                        # Set the corresponding row in 'ΤΥΠΟΣ' column to 'ΜΟΥΣΙΚΟ'
                        df.at[index, 'ΤΥΠΟΣ'] = 'ΜΟΥΣΙΚΟ'
                except Exception as e:
                    print(f"An error occurred during processing row {index}: {e}")

            # Add an else block to check if the column is empty and fill it with ''
            else:
                if df['ΤΥΠΟΣ'].empty:
                    df['ΤΥΠΟΣ'] = ''
            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def test_delete_mitronimo(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('ΜΗΤΡΩΝΥΜΟ', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_delete_mitronimo : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def test_delete_pinakas(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('ΠΙΝΑΚΑΣ', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_delete_pinakas : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def test_check_moria_pinaka(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')
            try:
                # Check and add 'ΜΟΡΙΑ ΠΙΝΑΚΑ' column if it doesn't exist
                check_if_column_exists(df, 'ΜΟΡΙΑ ΠΙΝΑΚΑ', None)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_check_moria_pinaka : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')


def test_normalize_perioxi_topothetisis(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            # Check and add 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column if it doesn't exist
            check_if_column_exists(df, 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ', '')
            try:

                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Α.Μ.Ω.', '')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Α.Π.Ω.', '')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('ΑΠΩ', '')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('ΑΜΩ', '')

            except Exception as e:
                print(f"An error occurred during preprocessing: {e}, {filename}")
            try:
                # Remove "B-" prefix from the 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Β-', '')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_normalize_perioxi_topothetisis(1) : {e},{filename}")
            try:
                # Remove the character "΄" from every row in the 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('΄', '')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_normalize_perioxi_topothetisis(1) : {e},{filename}")

            try:
                # Remove the character "(Π.Ε.)" from every row in the 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('(Π.Ε.)', '')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_normalize_perioxi_topothetisis(2) : {e},{filename}")

            try:
                # Remove the character "(Δ.Ε.)" from every row in the 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('(Δ.Ε.)', '')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('(Δ.Ε)', '')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_normalize_perioxi_topothetisis(3) : {e},{filename}")

            try:
                # Remove the character "(Δ.Ε.)" from every row in the 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Μειωμένου Ωραρίου', '')
            except Exception as e:
                print(f"An error occurred during preprocessing on method : {e},{filename}")

            try:
                # Remove the character "(Δ.Ε.)" from every row in the 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('-', '')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('  ', ' ')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_normalize_perioxi_topothetisis(4) : {e},{filename}")

            replacement_mapping = {

                "ΗΡΑΚΛΕΙΟΥ – Σχολείο Ευρωπαϊκής Παιδείας": "Ηρακλείου",

                "Μουσικό Σχολείο Λευκάδαςς": "Μουσικό Σχολείο Λευκάδας",
                "Μουσικό Σχολείο Ξάνθηςς": "Μουσικό Σχολείο Ξάνθης",
                "Μουσικό Σχολείο Πρέβεζαςς": "Μουσικό Σχολείο Πρέβεζας",

                "Ξάνθη – Μειονοτικά – ελληνόγλωσσο πρόγραμμα": "Ξάνθης",
                "ΞΑΝΘΗΣ – Μειονοτικά ελληνόγλωσσο πρόγραμμα": "Ξάνθης",

                "ΡΟΔΟΠΗΣ – Μειονοτικά ελληνόγλωσσο πρόγραμμα": "Ροδόπης",

                "Α ΕΒΡΟΥ – Μειονοτικά ελληνόγλωσσο πρόγραμμα": "Έβρου",

                "ΓΕΝ. ΕΚΚΛ/ΚΟ ΛΥΚΕΙΟΓΥΜΝΑΣΙΟ ΝΕΑΠΟΛΕΩΣ": "ΕΚΚΛ/ΚΟ ΛΥΚΕΙΟΓΥΜΝΑΣΙΟ",

                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΑΡΚΑΔΙΑΣ": "Αρκαδίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΗΡΑΚΛΕΙΟΥ": "Ηράκλειου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o ΠΕΙΡΑΙΑ": "Πειραιάς",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 1o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΣΕΡΡΩΝ": "Σερρών",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ 2o ΠΕΙΡΑΙΑ": "Πειραιάς",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ": "Αιτωλοακαρνανίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΑΡΓΟΛΙΔΑΣ": "Αργολίδας",

                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΑΡΤΑΣ": "Άρτας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΒΟΙΩΤΙΑΣ": "Βοιωτίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΓΡΕΒΕΝΩΝ": "Γρεβενών",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΔΡΑΜΑΣ": "Δράμας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ": "Δυτ. Αττικής",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΕΒΡΟΥ": "Έβρου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΕΥΒΟΙΑΣ": "Ευβοίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΕΥΡΥΤΑΝΙΑΣ": "Ευρυτανίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΖΑΚΥΝΘΟΥ": "Ζακύνθου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΗΛΕΙΑΣ": "Ηλείας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΗΜΑΘΙΑΣ": "Ημαθίας",

                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΘΕΣΠΡΩΤΙΑΣ": "Θεσπρωτίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΙΩΑΝΝΙΝΩΝ": "Ιωαννίνων",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΑΒΑΛΑΣ": "Καβάλας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΑΛΥΜΝΟΥ": "Καλύμνου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΑΡΔΙΤΣΑΣ": "Καρδίτσας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΑΣΤΟΡΙΑΣ": "Καστοριάς",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΕΡΚΥΡΑΣ": "Κέρκυρας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΕΦΑΛΛΗΝΙΑΣ": "Κεφαλληνίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΙΛΚΙΣ": "Κιλκίς",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΟΖΑΝΗΣ": "Κοζάνης",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΟΡΙΝΘΙΑΣ": "Κορινθίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΚΩ": "Κω",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΛΑΚΩΝΙΑΣ": "Λακωνίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΛΑΡΙΣΑΣ": "Λάρισας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΛΑΣΙΘΙΟΥ": "Λασιθίου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΛΕΣΒΟΥ": "Λέσβου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΛΕΥΚΑΔΑΣ": "Λευκάδας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΛΗΜΝΟΥ": "Λήμνου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΜΑΓΝΗΣΙΑΣ": "Μαγνησίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΜΕΣΣΗΝΙΑΣ": "Μεσσηνίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΝΑΞΟΥ": "Νάξου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΞΑΝΘΗΣ": "Ξάνθης",

                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΠΙΕΡΙΑΣ": "Πιερίας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΠΡΕΒΕΖΑΣ": "Πρέβεζας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΡΕΘΥΜΝΟΥ": "Ρεθύμνου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΡΟΔΟΠΗΣ": "Ροδόπης",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΡΟΔΟΥ": "Ρόδου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΣΑΜΟΥ": "Σάμου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΣΥΡΟΥ": "Σύρου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΤΡΙΚΑΛΩΝ": "Τρικάλων",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΦΘΙΩΤΙΔΑΣ": "Φθιώτιδας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΦΛΩΡΙΝΑΣ": "Φλώρινας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΦΩΚΙΔΑΣ": "Φωκίδας",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΧΑΛΚΙΔΙΚΗΣ": "Χαλκιδικής",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΧΑΝΙΩΝ": "Χανίων",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΧΙΟΥ": "Χίου",
                "ΣΔΕΥ στο ΚΕΔΑΣΥ ΠΕΛΛΑΣ": "Πέλλας",

                "ΚΕΔΑΣΥ 1o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",

                "ΣΔΕΥ ΚΕΔΑΣΥ 1o Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o ΒΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 1o ΠΕΙΡΑΙΑ": "Πειραιά",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "ΣΔΕΥ ΚΕΔΑΣΥ 2o ΠΕΙΡΑΙΑ": "Πειραιά",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ": "Αιτωλοακαρνανίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΑΡΓΟΛΙΔΑΣ": "Αργολίδας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΑΡΚΑΔΙΑΣ": "Αρκαδίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΑΡΤΑΣ": "Άρτας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΒΟΙΩΤΙΑΣ": "Βοιωτίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΓΡΕΒΕΝΩΝ": "Γρεβενών",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΔΡΑΜΑΣ": "Δράμας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ": "Δυτ. Αττικής",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΕΒΡΟΥ": "Έβρου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΕΥΒΟΙΑΣ": "Εύβοιας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΕΥΡΥΤΑΝΙΑΣ": "Ευρυτανίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΖΑΚΥΝΘΟΥ": "Ζακύνθου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΗΛΕΙΑΣ": "Ηλείας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΗΜΑΘΙΑΣ": "Ημαθίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΗΡΑΚΛΕΙΟΥ": "Ηρακλείου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΘΕΣΠΡΩΤΙΑΣ": "Θεσπρωτίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΙΩΑΝΝΙΝΩΝ": "Ιωαννίνων",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΑΒΑΛΑΣ": "Καβάλας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΑΛΥΜΝΟΥ": "Καλύμνου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΑΡΔΙΤΣΑΣ": "Καρδίτσας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΑΡΙΑΣ": "Καρίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΕΡΚΥΡΑΣ": "Κέρκυρας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΕΦΑΛΛΗΝΙΑΣ": "Κεφαλληνίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΙΛΚΙΣ": "Κιλκίς",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΟΖΑΝΗΣ": "Κοζάνης",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΟΡΙΝΘΙΑΣ": "Κορινθίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΚΩ": "Κω",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΛΑΚΩΝΙΑΣ": "Λακωνίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΛΑΡΙΣΑΣ": "Λαρίσας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΛΑΣΙΘΙΟΥ": "Λασιθίου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΛΕΣΒΟΥ": "Λέσβου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΛΕΥΚΑΔΑΣ": "Λευκάδας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΛΗΜΝΟΥ": "Λήμνου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΜΑΓΝΗΣΙΑΣ": "Μαγνησίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΜΕΣΣΗΝΙΑΣ": "Μεσσηνίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΝΑΞΟΥ": "Νάξου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΞΑΝΘΗΣ": "Ξάνθης",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΠΕΛΛΑΣ": "Πέλλας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΠΙΕΡΙΑΣ": "Πιερίας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΠΡΕΒΕΖΑΣ": "Πρέβεζας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΡΕΘΥΜΝΟΥ": "Ρεθύμνου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΡΟΔΟΠΗΣ": "Ροδόπης",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΡΟΔΟΥ": "Ρόδου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΣΑΜΟΥ": "Σάμου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΣΕΡΡΩΝ": "Σερρών",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΣΥΡΟΥ": "Σύρου",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΤΡΙΚΑΛΩΝ": "Τρικάλων",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΦΘΙΩΤΙΔΑΣ": "Φθιώτιδας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΦΛΩΡΙΝΑΣ": "Φλώρινας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΦΩΚΙΔΑΣ": "Φωκίδας",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΧΑΛΚΙΔΙΚΗΣ": "Χαλκιδικής",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΧΑΝΙΩΝ": "Χανίων",
                "ΣΔΕΥ ΚΕΔΑΣΥ ΧΙΟΥ": "Χίου",

                "ΚΕΣΥ 1o Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΚΕΣΥ 1o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΚΕΣΥ 1o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΚΕΣΥ 1o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΚΕΣΥ 1o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΚΕΣΥ 1o Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΚΕΣΥ 1o Β Θεσσαλονίκης": "Β Θεσσαλονίκης",
                "ΚΕΣΥ 1o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΚΕΣΥ 1o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "ΚΕΣΥ 1o ΠΕΙΡΑΙΑ": "Πειραιά",
                "ΚΕΣΥ 2o Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΚΕΣΥ 2o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΚΕΣΥ 2o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΚΕΣΥ 2o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΚΕΣΥ 2o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΚΕΣΥ 2o Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΚΕΣΥ 2o Β Θεσσαλονίκης": "Β Θεσσαλονίκης",
                "ΚΕΣΥ 2o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΚΕΣΥ 2o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "ΚΕΣΥ 2o ΠΕΙΡΑΙΑ": "Πειραιά",

                "ΚΕΣΥ Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΚΕΣΥ Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΣΔΕΥ Α Θεσσαλονίκης": "Α Θεσσαλονίκης",
                "ΚΕΣΥ ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΚΕΣΥ ΑΧΑΪΑΣ": "Αχαΐας",
                "ΚΕΣΥ Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΚΕΣΥ ΒΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΚΕΣΥ Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΚΕΣΥ ΠΕΙΡΑΙΑ": "Πειραιά",
                "ΚΕΣΥ ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ": "Αιτωλοακαρνανίας",
                "ΚΕΣΥ ΑΡΓΟΛΙΔΑΣ": "Αργολίδας",
                "ΚΕΣΥ ΑΡΚΑΔΙΑΣ": "Αρκαδίας",
                "ΚΕΣΥ ΑΡΤΑΣ": "Άρτας",
                "ΚΕΣΥ ΒΟΙΩΤΙΑΣ": "Βοιωτίας",
                "ΚΕΣΥ ΓΡΕΒΕΝΩΝ": "Γρεβενών",
                "ΚΕΣΥ ΔΡΑΜΑΣ": "Δράμας",
                "ΚΕΣΥ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ": "Δυτ. Αττικής",
                "ΚΕΣΥ ΕΒΡΟΥ": "Έβρου",
                "ΚΕΣΥ ΕΥΒΟΙΑΣ": "Εύβοιας",
                "ΚΕΣΥ ΕΥΡΥΤΑΝΙΑΣ": "Ευρυτανίας",
                "ΚΕΣΥ ΖΑΚΥΝΘΟΥ": "Ζακύνθου",
                "ΚΕΣΥ ΗΛΕΙΑΣ": "Ηλείας",
                "ΚΕΣΥ ΗΜΑΘΙΑΣ": "Ημαθίας",
                "ΚΕΣΥ ΗΡΑΚΛΕΙΟΥ": "Ηρακλείου",
                "ΚΕΣΥ ΘΕΣΠΡΩΤΙΑΣ": "Θεσπρωτίας",
                "ΚΕΣΥ ΙΩΑΝΝΙΝΩΝ": "Ιωαννίνων",
                "ΚΕΣΥ ΚΑΒΑΛΑΣ": "Καβάλας",
                "ΚΕΣΥ ΚΑΛΥΜΝΟΥ": "Καλύμνου",
                "ΚΕΣΥ ΚΑΡΔΙΤΣΑΣ": "Καρδίτσας",
                "ΚΕΣΥ ΚΑΣΤΟΡΙΑΣ": "Καστοριάς",
                "ΚΕΣΥ ΚΕΡΚΥΡΑΣ": "Κέρκυρας",
                "ΚΕΣΥ ΚΕΦΑΛΛΗΝΙΑΣ": "Κεφαλληνίας",
                "ΚΕΣΥ ΚΙΛΚΙΣ": "Κιλκίς",
                "ΚΕΣΥ ΚΟΖΑΝΗΣ": "Κοζάνης",
                "ΚΕΣΥ ΚΟΡΙΝΘΙΑΣ": "Κορινθίας",
                "ΚΕΣΥ ΚΩ": "Κω",
                "ΚΕΣΥ ΛΑΚΩΝΙΑΣ": "Λακωνίας",
                "ΚΕΣΥ ΛΑΡΙΣΑΣ": "Λάρισας",
                "ΚΕΣΥ ΛΑΣΙΘΙΟΥ": "Λασιθίου",

                "ΚΕΣΥ ΛΕΣΒΟΥ": "Λέσβου",
                "ΚΕΣΥ ΛΕΥΚΑΔΑΣ": "Λευκάδας",
                "ΚΕΣΥ ΛΗΜΝΟΥ": "Λήμνου",
                "ΚΕΣΥ ΜΑΓΝΗΣΙΑΣ": "Μαγνησίας",
                "ΚΕΣΥ ΜΕΣΣΗΝΙΑΣ": "Μεσσηνίας",
                "ΚΕΣΥ ΝΑΞΟΥ": "Νάξου",
                "ΚΕΣΥ ΞΑΝΘΗΣ": "Ξάνθης",
                "ΚΕΣΥ ΠΕΛΛΑΣ": "Πέλλας",
                "ΚΕΣΥ ΠΙΕΡΙΑΣ": "Πιερίας",
                "ΚΕΣΥ ΠΡΕΒΕΖΑΣ": "Πρέβεζας",
                "ΚΕΣΥ ΡΕΘΥΜΝΟΥ": "Ρεθύμνου",
                "ΚΕΣΥ ΡΟΔΟΠΗΣ": "Ροδόπης",
                "ΚΕΣΥ ΡΟΔΟΥ": "Ρόδου",
                "ΚΕΣΥ ΣΑΜΟΥ": "Σάμου",
                "ΚΕΣΥ ΣΕΡΡΩΝ": "Σερρών",
                "ΚΕΣΥ ΣΥΡΟΥ": "Σύρου",
                "ΚΕΣΥ ΤΡΙΚΑΛΩΝ": "Τρικάλων",
                "ΚΕΣΥ ΦΘΙΩΤΙΔΑΣ": "Φθιώτιδας",
                "ΚΕΣΥ ΦΛΩΡΙΝΑΣ": "Φλώρινας",
                "ΚΕΣΥ ΦΩΚΙΔΑΣ": "Φωκίδας",
                "ΚΕΣΥ ΧΑΛΚΙΔΙΚΗΣ": "Χαλκιδικής",
                "ΚΕΣΥ ΧΑΝΙΩΝ": "Χανίων",
                "ΚΕΣΥ ΧΙΟΥ": "Χίου",

                "ΣΔΕΥ στο Δ Αθήνας": "Δ Αθήνας",

                "ΣΔΕΥ στο Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΚΕΔΑΣΥ 1o Α ΑΘΗΝΑΣ": "Α Αθήνας",

                "ΚΕΔΑΣΥ 2o Α ΑΘΗΝΑΣ": "Α Αθήνας",
                "ΚΕΔΔΥ Β ΑΘΗΝΩΝ": "Β Αθήνας",
                "ΚΕΔΑΣΥ 1o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΚΕΔΑΣΥ 2o Β ΑΘΗΝΑΣ": "Β Αθήνας",
                "ΚΕΔΑΣΥ 1o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "ΚΕΔΑΣΥ 2o Γ ΑΘΗΝΑΣ": "Γ Αθήνας",
                "Κ.Ε.Σ.Υ. Γ ΑΘΗΝΩΝ": "Γ Αθήνας",
                "ΚΕΔΔΥ Γ ΑΘΗΝΩΝ": "Γ Αθήνας",
                "Κ.Ε.Σ.Υ. Δ ΑΘΗΝΩΝ": "Δ Αθήνας",
                "ΚΕΔΔΥ Δ ΑΘΗΝΩΝ": "Δ Αθήνας",
                "ΚΕΔΑΣΥ 1o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "ΚΕΔΑΣΥ 2o Δ ΑΘΗΝΑΣ": "Δ Αθήνας",
                "Κ.Ε.Σ.Υ. Α ΑΘΗΝΩΝ": "Α Αθήνας",
                "ΚΕΔΔΥ Α ΑΘΗΝΩΝ": "Α Αθήνας",
                "ΚΕΔΑΣΥ ΚΕΦΑΛΛΗΝΙΑΣ": "Κεφαλληνίας",
                "ΚΕΔΔΥ ΚΕΦΑΛΛΟΝΙΑ": "Κεφαλληνίας",
                "ΚΕΔΔΥ ΚΙΛΚΙΣ": "Κιλκίς",
                "ΚΕΔΑΣΥ ΚΙΛΚΙΣ": "Κιλκίς",
                "ΚΕΔΑΣΥ ΜΑΓΝΗΣΙΑΣ": "Μαγνησίας",
                "ΚΕΔΔΥ ΜΑΓΝΗΣΙΑΣ": "Μαγνησίας",
                "Κ.Ε.Σ.Υ. ΜΑΓΝΗΣΙΑΣ": "Μαγνησίας",
                "ΚΕΔΑΣΥ 1o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΚΕΔΑΣΥ 2o ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",
                "ΚΕΔΔΥ ΑΝΑΤ. ΑΤΤΙΚΗ": "Ανατ. Αττικής",
                "Κ.Ε.Σ.Υ. ΔΥΤ. ΑΤΤΙΚΗ": "Α Δυτ. Αττικής",
                "ΚΕΔΔΥ ΔΥΤ. ΑΤΤΙΚΗ": "Α Δυτ. Αττικής",
                "ΚΕΔΔΥ ΣΑΜΟΥ": "Σάμου",
                "ΚΕΔΑΣΥ ΣΑΜΟΥ": "Σάμου",
                "ΚΕΔΔΥ ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑ": "Αιτωλοακαρνανίας",
                "ΚΕΔΑΣΥ ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ": "Αιτωλοακαρνανίας",
                "ΚΕΔΔΥ ΑΡΓΟΛΙΔΑ": "Αργολίδας",
                "ΚΕΔΑΣΥ ΑΡΓΟΛΙΔΑΣ": "Αργολίδας",
                "ΚΕΔΔΥ ΑΡΚΑΔΙΑ": "Αρκαδίας",
                "ΚΕΔΑΣΥ ΑΡΚΑΔΙΑΣ": "Αρκαδίας",
                "ΚΕΔΔΥ ΒΟΙΩΤΙΑ": "Βοιωτίας",
                "ΚΕΔΑΣΥ ΒΟΙΩΤΙΑΣ": "Βοιωτίας",
                "ΚΕΔΑΣΥ ΓΡΕΒΕΝΩΝ": "Γρεβενών",
                "Κ.Ε.Σ.Υ. ΓΡΕΒΕΝΑ": "Γρεβενών",
                "ΚΕΔΔΥ ΓΡΕΒΕΝΑ": "Γρεβενών",
                "ΚΕΔΔΥ Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",

                "ΚΕΔΑΣΥ 2o Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "Κ.Ε.Σ.Υ. Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",
                "ΚΕΔΑΣΥ 1o Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΚΕΔΑΣΥ 1o ΒΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΚΕΔΑΣΥ 2o Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",
                "ΚΕΔΑΣΥ 1o ΠΕΙΡΑΙΑ": "Πειραιά",
                "ΚΕΔΑΣΥ 2o ΠΕΙΡΑΙΑ": "Πειραιά",
                "ΚΕΔΔΥ ΕΒΡΟΣ": "Έβρου",
                "ΚΕΔΑΣΥ ΕΒΡΟΥ": "Έβρου",
                "ΚΕΔΑΣΥ ΙΩΑΝΝΙΝΩΝ": "Ιωαννίνων",
                "ΚΕΔΔΥ ΙΩΑΝΝΙΝΩΝ": "Ιωαννίνων",

                "ΚΕΔΑΣΥ ΚΑΒΑΛΑΣ": "Καβάλας",
                "ΚΕΔΔΥ ΚΑΒΑΛΑΣ": "Καβάλας",
                "ΚΕΔΑΣΥ ΚΟΡΙΝΘΙΑΣ": "Κορινθίας",
                "ΚΕΔΑΣΥ ΛΑΣΙΘΙΟΥ": "Λασιθίου",
                "Κ.Ε.Σ.Υ. ΛΑΣΙΘΙ": "Λασιθίου",
                "ΚΕΔΔΥ ΑΡΤΑ": "Άρτας",
                "ΚΕΔΑΣΥ ΑΡΤΑΣ": "Άρτας",
                "ΚΕΔΔΥ ΑΧΑΙΑ": "Αχαΐας",
                "ΚΕΔΑΣΥ 1o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΚΕΔΑΣΥ 2o ΑΧΑΪΑΣ": "Αχαΐας",
                "ΚΕΔΑΣΥ ΧΑΛΚΙΔΙΚΗΣ": "Χαλκιδικής",
                "ΚΕΔΔΥ ΧΑΛΚΙΔΙΚΗ": "Χαλκιδικής",
                "ΚΕΔΑΣΥ ΧΑΝΙΩΝ": "Χανίων",
                "ΚΕΔΔΥ ΧΑΝΙΑ": "Χανίων",
                "ΚΕΔΔΥ ΛΕΣΒΟΥ": "Γ Λέσβου",
                "ΚΕΔΑΣΥ ΛΕΣΒΟΥ": "Γ Λέσβου",
                "ΚΕΔΑΣΥ ΖΑΚΥΝΘΟΥ": "Ζακύνθου",
                "ΚΕΔΔΥ ΖΑΚΥΝΘΟΣ": "Ζακύνθου",
                "ΚΕΔΔΥ ΗΛΕΙΑ": "Ηλείας",
                "ΚΕΔΑΣΥ ΗΛΕΙΑΣ": "Ηλείας",
                "ΚΕΔΑΣΥ ΗΜΑΘΙΑΣ": "Ημαθίας",
                "ΚΕΔΔΥ ΗΜΑΘΙΑ": "Ημαθίας",
                "ΚΕΔΔΥ ΗΡΑΚΛΕΙΟ": "Ηρακλείου",
                "ΚΕΔΑΣΥ ΗΡΑΚΛΕΙΟΥ": "Ηρακλείου",
                "ΚΕΔΔΥ ΘΕΣΠΡΩΤΙΑ": "Θεσπρωτίας",
                "ΚΕΔΑΣΥ ΘΕΣΠΡΩΤΙΑΣ": "Θεσπρωτίας",
                "ΚΕΔΔΥ ΚΑΡΔΙΤΣΑ": "Καρδίτσας",
                "ΚΕΔΑΣΥ ΚΑΡΔΙΤΣΑΣ": "Καρδίτσας",
                "Κ.Ε.Σ.Υ. ΚΑΡΔΙΤΣΑ": "Καρδίτσας",
                "ΚΕΔΔΥ ΚΑΣΤΟΡΙΑ": "Καστοριάς",
                "ΚΕΔΑΣΥ ΚΑΣΤΟΡΙΑΣ": "Καστοριάς",
                "ΚΕΔΑΣΥ ΚΑΡΙΑΣ": "Καρίας",
                "ΚΕΔΔΥ ΚΟΖΑΝΗ": "Κοζάνης",
                "ΚΕΔΑΣΥ ΚΟΖΑΝΗΣ": "Κοζάνης",
                "ΚΕΔΑΣΥ ΛΑΚΩΝΙΑΣ": "Λακωνίας",
                "ΚΕΔΔΥ ΛΑΚΩΝΙΑ": "Λακωνίας",
                "ΚΕΔΔΥ ΛΕΥΚΑΔΑ": "Λευκάδας",
                "ΚΕΔΑΣΥ ΛΕΥΚΑΔΑΣ": "Λευκάδας",
                "ΚΕΔΔΥ ΜΕΣΣΗΝΙΑ": "Μεσσηνίας",
                "ΚΕΔΑΣΥ ΜΕΣΣΗΝΙΑΣ": "Μεσσηνίας",
                "ΚΕΔΑΣΥ ΠΕΛΛΑΣ": "Πέλλας",
                "ΚΕΔΔΥ ΠΕΛΛΑ": "Πέλλας",
                "ΚΕΔΑΣΥ ΞΑΝΘΗΣ": "Ξάνθης",
                "ΚΕΔΔΥ ΞΑΝΘΗ": "Ξάνθης",
                "ΚΕΔΔΥ ΠΙΕΡΙΑ": "Πιερίας",
                "ΚΕΔΑΣΥ ΠΙΕΡΙΑΣ": "Πιερίας",
                "Κ.Ε.Σ.Υ. ΠΕΙΡΑΙΑΣ": "Πιερίας",
                "Κ.Ε.Σ.Υ. ΠΕΙΡΑΙΑ": "Πιερίας",
                "ΚΕΔΔΥ ΠΕΙΡΑΙΑΣ": "Πιερίας",
                "ΚΕΔΔΥ ΡΟΔΟΠΗ": "Ροδόπης",
                "ΚΕΔΑΣΥ ΡΟΔΟΠΗΣ": "Ροδόπης",
                "ΚΕΔΑΣΥ ΠΡΕΒΕΖΑΣ": "Πρέβεζας",
                "ΚΕΔΔΥ ΠΡΕΒΕΖΑ": "Πρέβεζας",
                "ΚΕΔΑΣΥ ΤΡΙΚΑΛΩΝ": "Τρικάλων",
                "ΚΕΔΔΥ ΤΡΙΚΑΛΑ": "Τρικάλων",
                "ΚΕΔΑΣΥ ΣΕΡΡΩΝ": "Σερρών",
                "ΚΕΔΔΥ ΣΕΡΡΕΣ": "Σερρών",
                "ΚΕΔΑΣΥ ΦΘΙΩΤΙΔΑΣ": "Φθιώτιδας",

                "ΚΕΔΔΥ ΦΘΙΩΤΙΔΑ": "Φθιώτιδας",
                "ΚΕΔΔΥ ΦΛΩΡΙΝΑ": "Φλώρινας",
                "ΚΕΔΑΣΥ ΦΛΩΡΙΝΑΣ": "Φλώρινας",
                "ΚΕΔΔΥ ΚΩΣ": "Κω",
                "ΚΕΔΑΣΥ ΚΩ": "Κω",
                "ΚΕΔΔΥ ΛΑΡΙΣΑΣ": "Λάρισας",
                "ΚΕΔΑΣΥ ΛΑΡΙΣΑΣ": "Λάρισας",
                "ΚΕΔΔΥ ΚΕΡΚΥΡΑ": "Κέρκυρας",
                "ΚΕΔΑΣΥ ΚΕΡΚΥΡΑΣ": "Κέρκυρας",
                "ΚΕΔΔΥ ΛΗΜΝΟΥ": "Λήμνου",
                "ΚΕΔΑΣΥ ΛΗΜΝΟΥ": "Λήμνου",
                "ΚΕΔΔΥ ΝΑΞΟΥ": "Νάξου",
                "ΚΕΔΑΣΥ ΝΑΞΟΥ": "Νάξου",
                "ΚΕΔΔΥ ΡΟΔΟΥ": "Ρόδου",
                "ΚΕΔΑΣΥ ΡΟΔΟΥ": "Ρόδου",
                "ΚΕΔΔΥ ΣΥΡΟΥ": "Σύρου",
                "ΚΕΔΑΣΥ ΣΥΡΟΥ": "Σύρου",
                "ΚΕΔΔΥ ΦΩΚΙΔΑ": "Φωκίδας",
                "ΚΕΔΑΣΥ ΦΩΚΙΔΑΣ": "Φωκίδας",
                "ΚΕΔΑΣΥ ΧΙΟΥ": "Χίου",
                "ΚΕΔΔΥ ΧΙΟΣ": "Χίου",
                "ΚΕΔΔΥ ΚΑΛΥΜΝΟΥ": "Καλύμνου",
                "ΚΕΔΑΣΥ ΚΑΛΥΜΝΟΥ": "Καλύμνου",
                "ΚΕΔΑΣΥ ΕΥΒΟΙΑΣ": "Εύβοιας",
                "ΚΕΔΔΥ ΕΥΒΟΙΑ": "Εύβοιας",
                "ΚΕΔΑΣΥ ΕΥΡΥΤΑΝΙΑΣ": "Ευρυτανίας",
                "ΚΕΔΔΥ ΕΥΡΥΤΑΝΙΑ": "Ευρυτανίας",
                "ΚΕΔΔΥ ΔΡΑΜΑ": "Δράμας",
                "ΚΕΔΑΣΥ ΔΡΑΜΑΣ": "Δράμας",
                "ΚΕΔΑΣΥ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ": "Δυτ. Αττικής",

                "ΚΕΔΑΣΥ ΡΕΘΥΜΝΟΥ": "Ρεθύμνου",
                "ΚΕΔΔΥ ΡΕΘΥΜΝΟ": "Ρεθύμνου",
                "Κ.Ε.Σ.Υ. ΠΙΕΡΙΑ": "Πιερίας",

                "Α Αθηνών": "Α Αθήνας",
                "Α ΑΘΗΝΑΣ": "Α Αθήνας",

                "ΑΑ Αθηνών": "Α Αθήνας",
                "ΑΑ Αθήνας": "Α Αθήνας",
                "Β Αθηνών": "Β Αθήνας",

                "Β ΑΘΗΝΑΣ": "Β Αθήνας",

                "Γ ΑΘΗΝΑΣ": "Γ Αθήνας",

                "Δ ΑΘΗΝΑΣ": "Δ Αθήνας",

                "Α Κεφαλλονιά": "Α Κεφαλληνίας",
                "ΑΑ Κεφαλληνίας": "Α Κεφαλληνίας",
                "Β ΚΕΦΑΛΛΗΝΙΑΣ": "Β Κεφαλληνίας",
                "ΚΕΦΑΛΛΗΝΙΑΣ": "Κεφαλληνίας",
                "Β Κεφαλλονιά": "Β Κεφαλληνίας",

                "ΑΑ Κεφαλλονιά": "Α Κεφαλληνίας",
                "ΚΙΛΚΙΣ": "Κιλκίς",

                "Α ΜΑΓΝΗΣΙΑΣ(Δ.Ε)": "Α Μαγνησίας",
                "ΜΑΓΝΗΣΙΑΣ": "Μαγνησίας",

                "ΑΑ Ανατ. Αττικής": "Α Ανατ. Αττικής",
                "ΑΒ Ανατ. Αττικής": "Β Ανατ. Αττικής",
                "ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ": "Ανατ. Αττικής",

                "ΑΔυτ. Αττικής": "Α Δυτ. Αττικής",

                "ΑΑ Λέσβου": "Α Λέσβου",
                "ΑΑ Πειραιά": "Α Πειραιά",
                "ΑΑ Σάμου": "Α Σάμου",
                "ΣΑΜΟΥ": "Σάμου",

                "ΑΑ Χίου": "Α Χίου",
                "ΑΑιτωλοακαρνανία": "Α Αιτωλοακαρνανίας",
                "Αιτωλ/νίας": "Αιτωλοακαρνανίας",
                "Αιτωλοακαρνανία": "Αιτωλοακαρνανίας",
                "ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ": "Αιτωλοακαρνανίας",
                "Α Αιτωλοακαρνανίαςς": "Α Αιτωλοακαρνανίας",
                "Αιτωλοακαρνανίαςς": "Αιτωλοακαρνανίας",

                "ΑΡΓΟΛΙΔΑΣ": "Αργολίδας",
                "Α Αργολίδαςς": "Α Αργολίδας",

                "ΑΡΚΑΔΙΑΣ": "Αρκαδίας",

                "ΑΑχαΐα": "Α Αχαΐας",

                "Αργολίδαςς": "Αργολίδας",

                "ΒΟΙΩΤΙΑΣ": "Βοιωτίας",

                "ΓΡΕΒΕΝΩΝ": "Γρεβενών",

                "ΑΒ Θεσσαλονίκης": "Β Θεσσαλονίκης",
                "Α ΘΕΣΣΑΛΟΝΙΚΗΣ": "Α Θεσσαλονίκης",

                "Β ΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",

                "ΒΘΕΣΣΑΛΟΝΙΚΗΣ": "Β Θεσσαλονίκης",

                "ΑΒ Πειραιά": "Β Πειραιά",
                "ΠΕΙΡΑΙΑ": "Πειραιά",

                "ΑΒοιωτία": "Α Βοιωτίας",
                "ΑΓ Αθηνών": "Γ Αθήνας",
                "ΑΔ Αθηνών": "Δ Αθήνας",
                "ΑΔράμα": "Α Δράμας",
                "ΑΔωδεκάνησα": "Α Δωδεκανήσου",
                "ΑΕβρος": "Α Έβρου",

                "ΕΒΡΟΥ": "Έβρου",

                "ΑΕύβοια": "Α Ευβοίας",
                "ΑΖάκυνθος": "Α Ζακύνθου",
                "ΑΗλεία": "Α Ηλείας",
                "ΑΗράκλειο": "Α Ηρακλείου",
                "Β ΣΑΜΟΥ": "Β Σάμου",
                "ΑΑργολίδα": "Α Αργολίδας",
                "Β Γ Λέσβου": "Γ Λέσβου",
                "Α Γ Λέσβου": "Γ Λέσβου",
                "ΑΙωαννίνων": "Α Ιωαννίνων",
                "ΙΩΑΝΝΙΝΩΝ": "Ιωαννίνων",

                "ΑΚαβάλας": "Α Καβάλας",
                "ΚΑΒΑΛΑΣ": "Καβάλας",

                "ΑΚαστοριά": "Α Καστοριάς",
                "ΑΚέρκυρα": "Α Κέρκυρας",
                "ΑΚιλκίς": "Α Κιλκίς",
                "ΑΚορινθία": "Α Κορινθίας",
                "Κορινθία": "Κορινθίας",
                "Α Κορινθίαςς": "Κορινθίας",
                "Κορινθίαςς": "Κορινθίας",
                "ΚΟΡΙΝΘΙΑΣ": "Κορινθίας",

                "ΑΚυκλάδες": "Α Κυκλάδων",
                "Β Κυκλάδων": "Β Κυκλάδων",
                "Β ΚΥΚΛΑΔΩΝ": "Β Κυκλάδων",
                "Α ΚΥΚΛΑΔΩΝ": "Α Κυκλάδων",
                "ΑΛασίθι": "Α Λασιθίου",
                "Λασίθι": "Λασιθίου",
                "ΛΑΣΙΘΙΟΥ": "Λασιθίου",

                "ΛΑΣΙΘΙ": "Λασιθίου",

                "ΑΠιερία": "Α Πιερίας",
                "ΑΠρέβεζα": "Α Πρέβεζας",
                "Αργολίδα": "Αργολίδας",
                "Αρκαδία": "Αρκαδίας",
                "Αρκαδίαςς": "Αρκαδίας",
                "Αρτα": "Άρτας",
                "Άρτα": "Άρτας",

                "Άρταςς": "Άρτας",
                "ΑΡΤΑΣ": "Άρτας",

                "ΑΦωκίδα": "Α Φωκίδας",
                "Α Φωκίδαςς": "Α Φωκίδας",
                "Αχαΐα": "Αχαΐας",
                "ΑΧΑΪΑΣ": "Αχαΐας",

                "Αχαΐαςς": "Αχαΐας",
                "ΑΧαλκιδική": "Α Χαλκιδικής",
                "Α Χαλκιδικήςς": "Α Χαλκιδικής",
                "ΧΑΛΚΙΔΙΚΗΣ": "Χαλκιδικής",
                "Χαλκιδικήςς": "Χαλκιδικής",

                "ΑΧανιά": "Α Χανίων",
                "ΧΑΝΙΩΝ": "Χανίων",

                "Βοιωτία": "Βοιωτίας",
                "Βοιωτίαςς": "Βοιωτίας",
                "Α Βοιωτίαςς": "Α Βοιωτίας",
                "Γ Αθηνών": "Γ Αθήνας",
                "Γ Δωδεκανήσου": "Γ Δωδεκανήσου",
                "Γ Κυκλάδων": "Γ Κυκλάδων",
                "Γ ΛΕΣΒΟΥ": "Γ Λέσβου",
                "ΛΕΣΒΟΥ": "Λέσβου",

                "Γρεβενά": "Γρεβενών",
                "Δ Αθηνών": "Δ Αθήνας",
                "Δράμα": "Δράμας",
                "Δράμαςς": "Δράμας",
                "Εβρος": "Έβρου",
                "Ευρυτανία": "Ευρυτανίας",
                "Ευρυτανίαςς": "Ευρυτανίας",
                "Ζάκυνθος": "Ζακύνθου",
                "ΖΑΚΥΝΘΟΥ": "Ζακύνθου",

                "Ηλεία": "Ηλείας",
                "ΗΛΕΙΑΣ": "Ηλείας",

                "Ηλείαςς": "Ηλείας",
                "Ημαθία": "Ημαθίας",
                "Ημαθίαςς": "Ημαθίας",
                "ΗΜΑΘΙΑΣ": "Ημαθίας",

                "Ηράκλειο": "Ηρακλείου",
                "Ηρακλείουυ": "Ηρακλείου",
                "ΗΡΑΚΛΕΙΟΥ": "Ηρακλείου",

                "Θεσπρωτία": "Θεσπρωτίας",
                "Θεσπρωτίαςς": "Θεσπρωτίας",
                "ΘΕΣΠΡΩΤΙΑΣ": "Θεσπρωτίας",

                "Καρδίτσα": "Καρδίτσας",
                "Καρδίτσαςς": "Καρδίτσας",
                "ΚΑΡΔΙΤΣΑΣ": "Καρδίτσας",

                "Πειραιά": "Πειραιάς",

                "Καστοριά": "Καστοριάς",
                "Καστοριάςς": "Καστοριάς",
                "ΚΑΣΤΟΡΙΑΣ": "Καστοριάς",

                "Γ ΔΩΔΕΚΑΝΗΣΟΥ": "Γ Δωδεκανήσου",
                "Β ΔΩΔΕΚΑΝΗΣΟΥ": "Β Δωδεκανήσου",
                "Α ΔΩΔΕΚΑΝΗΣΟΥ": "Α Δωδεκανήσου",
                "Γ ΚΥΚΛΑΔΩΝ": "Γ Κυκλάδων",
                "Κοζάνη": "Κοζάνης",
                "ΚΟΖΑΝΗΣ": "Κοζάνης",
                "Κοζάνηςς": "Κοζάνης",

                "Λακωνία": "Λακωνίας",
                "ΛΑΚΩΝΙΑΣ": "Λακωνίας",
                "Λακωνίαςς": "Λακωνίας",

                "Λευκάδα": "Λευκάδας",
                "ΛΕΥΚΑΔΑΣ": "Λευκάδας",
                "Λευκάδαςς": "Λευκάδας",

                "Μεσσηνία": "Μεσσηνίας",
                "ΜΕΣΣΗΝΙΑΣ": "Μεσσηνίας",
                "Μεσσηνίαςς": "Μεσσηνίας",

                "Ξάνθη": "Ξάνθης",
                "ΞΑΝΘΗΣ": "Ξάνθης",
                "Ξάνθηςς": "Ξάνθης",

                "Πέλλα": "Πέλλας",
                "ΠΕΛΛΑΣ": "Πέλλας",

                "Πέλλαςς": "Πέλλας",

                "Πιερία": "Πιερίας",
                "Α Πιερίαςς": "Α Πιερίας",
                "Πιερίαςς": "Πιερίας",
                "ΠΙΕΡΙΑΣ": "Πιερίας",

                "Πρέβεζα": "Πρέβεζας",
                "ΠΡΕΒΕΖΑΣ": "Πρέβεζας",
                "Α Πρέβεζαςς": "Α Πρέβεζας",
                "Πρέβεζαςς": "Πρέβεζας",

                "Ρέθυμνο": "Ρεθύμνου",
                "ΡΕΘΥΜΝΟΥ": "Ρεθύμνου",

                "Ροδόπη": "Ροδόπης",
                "ΡΟΔΟΠΗΣ": "Ροδόπης",
                "Ροδόπηςς": "Ροδόπης",

                "Α ΜΑΓΝΗΣΙΑΣ (Δ.Ε)": "Α Μαγνησίας",
                "Τρίκαλα": "Τρικάλων",
                "ΤΡΙΚΑΛΩΝ": "Τρικάλων",

                "Σέρρες": "Σερρών",
                "ΣΕΡΡΩΝ": "Σερρών",
                "ΚΑΡΙΑΣ": "Κάριας",

                "ΦΘΙΩΤΙΔΑΣ": "Φθιώτιδας",

                "Φθιώτιδος": "Φθιώτιδας",
                "Φλώρινα": "Φλώρινας",
                "ΦΛΩΡΙΝΑΣ": "Φλώρινας",

                "Φωκίδα": "Φωκίδας",
                "Χαλκιδική": "Χαλκιδικής",
                "Χανιά": "Χανίων",
                "ΚΩ": "Κω",

                "ΛΑΡΙΣΑΣ": "Λάρισας",

                "ΚΕΡΚΥΡΑΣ": "Κέρκυρας",
                "Κέρκυρα": "Κέρκυρας",

                "ΛΗΜΝΟΥ": "Λήμνου",

                "ΝΑΞΟΥ": "Νάξου",

                "ΡΟΔΟΥ": "Ρόδου",

                "ΣΥΡΟΥ": "Σύρου",

                "ΦΩΚΙΔΑΣ": "Φωκίδας",

                "ΧΙΟΥ": "Χίου",

                "ΚΑΛΥΜΝΟΥ": "Καλύμνου",
                "ΕΥΒΟΙΑΣ": "Εύβοιας",
                "Εύβοια": "Εύβοιας",
                "ΕΥΡΥΤΑΝΙΑΣ": "Ευρυτανίας",
                "ΔΡΑΜΑΣ": "Δράμας",
                "ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ": "Δυτ. Αττικής",

                # "": "",
            }

            try:
                for old_string, new_string in replacement_mapping.items():
                    df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(old_string, new_string,
                                                                                      regex=True)
                # df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].map(replacement_mapping).fillna(df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'])
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    "Ειδικό Αγροτικό Κατάστημα Κράτησης Νέων Κασσαβέτειας [ΠΑΡΑΡΤΗΜΑ ΓΥΜΝΑΣΙΟΥ Ν. ΑΓΧΙΑΛΟΥ (Ε.Α.Κ.Κ.Ν.Κ.)]",
                    'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Ειδικό Αγροτικό Σωφρονιστικό Κατάστημα Νέων Κασσαβέτειας [ΠΑΡΑΡΤΗΜΑ ΓΥΜΝΑΣΙΟΥ Ν. ΑΓΧΙΑΛΟΥ (Ε.Α.Σ.Κ.Ν.Κ.)]',
                    'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Ειδικό Κατάστημα Κράτησης Νέων Αυλώνα [2ο ΓΥΜΝΑΣΙΟ ΑΥΛΩΝΑ ΜΕ ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ (Ε.Κ.Κ.Ν.Α.)]',
                    'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Ειδικό Κατάστημα Κράτησης Νέων Βόλου [ΠΑΡΑΡΤΗΜΑ 6ου ΓΥΜΝΑΣΙΟΥ ΒΟΛΟΥ (Ε.Κ.Κ.Ν.Β.)]', 'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Ειδικό Σωφρονιστικό Κατάστημα Νέων Αυλώνα [2ο ΓΥΜΝΑΣΙΟ ΑΥΛΩΝΑ ΜΕ ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ (Ε.Σ.Κ.Ν.Α.)]',
                    'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Ειδικό Σωφρονιστικό Κατάστημα Νέων Αυλώνα [2ο Τριθέσιο Δημοτικό Σχολείο Αυλώνας]', 'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Ειδικό Σωφρονιστικό Κατάστημα Νέων Βόλου [ΠΑΡΑΡΤΗΜΑ 6ου ΓΥΜΝΑΣΙΟΥ ΒΟΛΟΥ (Ε.Σ.Κ.Ν.Β.)]',
                    'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Σωφρονιστικό Κατάστημα Γρεβενών [Μονοθέσιο Δημοτικό Σχολείο]', 'Κατάστημα')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace(
                    'Κατάστημα Κράτησης Κορίνθου [ΠΑΡΑΡΤΗΜΑ ΕΣΠΕΡΙΝΟΥ ΓΥΜΝΑΣΙΟΥ ΚΟΡΙΝΘΟΥ ΚΑΙ ΠΑΡΑΡΤΗΜΑ 1ου ΓΕΛ ΚΟΡΙΝΘΟΥ (Κ.Κ.Κ.)]',
                    'Κατάστημα')

                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('ΚΕΣΥ 1o Β Θεσσαλονίκης',
                                                                                  'Β Θεσσαλονίκης')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('ΚΕΣΥ 2o Β Θεσσαλονίκης',
                                                                                  'Β Θεσσαλονίκης')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Αργολίδαςς', 'Αργολίδας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Άρταςς', 'Άρτας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Α Αργολίδαςς', 'Α Αργολίδας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Μουσικό Σχολείο Αργολίδαςς',
                                                                                  'Μουσικό Σχολείο Αργολίδας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Μουσικό Σχολείο Άρταςς',
                                                                                  'Μουσικό Σχολείο Άρτας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Μουσικό Σχολείο Κέρκυραςς',
                                                                                  'Μουσικό Σχολείο Κέρκυρας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Φθιώτιδα', 'Φθιώτιδας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Φθιώτιδαςς', 'Φθιώτιδας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Φλώριναςς', 'Φλώρινας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Φωκίδαςς', 'Φωκίδας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Χαλκιδικήςς', 'Χαλκιδικής')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Πειραιάςς', 'Πειραιάς')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Α Κέρκυραςς', 'Α Κέρκυρας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Κέρκυραςς', 'Κέρκυρας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Α Εύβοιαςς', 'Α Εύβοιας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Εύβοιαςς', 'Εύβοιας')
                df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.replace('Β ΑΝΑΤ. ΑΤΤΙΚΗΣ', 'Β Ανατ. Αττικής')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_normalize_perioxi_topothetisis(5) : {e},{filename}")

                # Remove key words from the 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def normalize_klados_values(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            # Check and add 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' column if it doesn't exist
            check_if_column_exists(df, 'ΚΛΑΔΟΣ', '')

            replacements = {
                'ΠΕ01.00': 'ΠΕ01',
                'ΠΕ02.00': 'ΠΕ02',
                'ΠΕ03.00': 'ΠΕ03',
                'ΠΕ05.00': 'ΠΕ05',
                'ΠΕ06.00': 'ΠΕ06',
                'ΠΕ07.00': 'ΠΕ07',
                'ΠΕ08.00': 'ΠΕ08',
                'ΠΕ11.00': 'ΠΕ11',
                'ΠΕ80.00': 'ΠΕ80',
                'ΠΕ81.00': 'ΠΕ81',
                'ΠΕ82.00': 'ΠΕ82',
                'ΠΕ83.00': 'ΠΕ83',
                'ΠΕ84.00': 'ΠΕ84',
                'ΠΕ86.00': 'ΠΕ86',
                'ΠΕ90.00': 'ΠΕ90',
                'ΤΕ16.00': 'ΤΕ16',
                'ΠΕ10.00': 'ΠΕ78',
                'ΠΕ13.00': 'ΠΕ78',
                'ΠΕ16.01': 'ΠΕ79.01',
                'ΠΕ16.02': 'ΠΕ79.01',
                'ΠΕ17.14': 'ΠΕ79.02',
                'ΠΕ17.13': 'ΠΕ79.02',
                'ΠΕ09.00': 'ΠΕ80',
                'ΠΕ18.02': 'ΠΕ80',
                'ΠΕ18.03': 'ΠΕ80',
                'ΠΕ18.35': 'ΠΕ80',
                'ΠΕ18.40': 'ΠΕ80',
                'ΠΕ15.00': 'ΠΕ80',
                'ΠΕ12.01': 'ΠΕ81',
                'ΠΕ12.02': 'ΠΕ81',
                'ΠΕ12.03': 'ΠΕ81',
                'ΠΕ17.01': 'ΠΕ81',
                'ΠΕ17.05': 'ΠΕ81',
                'ΠΕ17.11': 'ΠΕ81',
                'ΠΕ12.04': 'ΠΕ82',
                'ΠΕ12.07': 'ΠΕ82',
                'ΠΕ12.11': 'ΠΕ82',
                'ΠΕ17.02': 'ΠΕ82',
                'ΠΕ17.06': 'ΠΕ82',
                'ΠΕ18.18': 'ΠΕ82',
                'ΠΕ18.31': 'ΠΕ82',
                'ΠΕ18.32': 'ΠΕ82',
                'ΠΕ12.05': 'ΠΕ83',
                'ΠΕ17.03': 'ΠΕ83',
                'ΠΕ17.07': 'ΠΕ83',
                'ΠΕ12.06': 'ΠΕ84',
                'ΠΕ17.04': 'ΠΕ84',
                'ΠΕ17.08': 'ΠΕ84',
                'ΠΕ17.09': 'ΠΕ84',
                'ΠΕ12.08': 'ΠΕ85',
                'ΠΕ19.00': 'ΠΕ86',
                'ΠΕ20.00': 'ΠΕ86',
                'ΠΕ14.01': 'ΠΕ87.01',
                'ΠΕ14.02': 'ΠΕ87.01',
                'ΠΕ14.03': 'ΠΕ87.01',
                'ΠΕ14.06': 'ΠΕ87.02',
                'ΠΕ18.10': 'ΠΕ87.02',
                'ΠΕ18.11': 'ΠΕ87.02',
                'ΠΕ18.39': 'ΠΕ87.02',
                'ΠΕ18.07': 'ΠΕ87.04',
                'ΠΕ18.08': 'ΠΕ87.05',
                'ΠΕ18.09': 'ΠΕ87.06',
                'ΠΕ18.21': 'ΠΕ87.07',
                'ΠΕ18.25': 'ΠΕ87.08',
                'ΠΕ18.24': 'ΠΕ87.08',
                'ΠΕ18.33': 'ΠΕ87.09',
                'ΠΕ18.37': 'ΠΕ87.10',
                'ΠΕ14.04': 'ΠΕ88.01',
                'ΠΕ18.12': 'ΠΕ88.02',
                'ΠΕ18.15': 'ΠΕ88.02',
                'ΠΕ18.17': 'ΠΕ88.02',
                'ΠΕ18.30': 'ΠΕ88.02',
                'ΠΕ18.13': 'ΠΕ88.03',
                'ΠΕ18.14': 'ΠΕ88.03',
                'ΠΕ18.36': 'ΠΕ88.04',
                'ΠΕ18.01': 'ΠΕ89.01',
                'ΠΕ18.26': 'ΠΕ89.01',
                'ΠΕ18.27': 'ΠΕ89.01',
                'ΠΕ18.28': 'ΠΕ89.01',
                'ΠΕ18.29': 'ΠΕ89.01',
                'ΠΕ18.23': 'ΠΕ90',
                'ΠΕ32.00': 'ΠΕ91.01',
                'ΠΕ18.41': 'ΠΕ91.02',
                'ΤΕ01.01': 'ΤΕ02.01',
                'ΤΕ01.05': 'ΤΕ02.01',
                'ΤΕ01.02': 'ΤΕ02.02',
                'ΤΕ01.03': 'ΤΕ02.02',
                'ΤΕ01.28': 'ΤΕ02.02',
                'ΤΕ01.08': 'ΤΕ02.03',
                'ΤΕ01.23': 'ΤΕ02.03',
                'ΤΕ01.10': 'ΤΕ02.04',
                'ΤΕ01.11': 'ΤΕ02.04',
                'ΤΕ01.12': 'ΤΕ02.05',
                'ΤΕ01.14': 'ΤΕ02.05',
                'ΤΕ01.17': 'ΤΕ02.05',
                'ΤΕ01.22': 'ΤΕ02.06',
                'ΤΕ01.27': 'ΤΕ02.06',
                'ΤΕ01.32': 'ΤΕ02.07',
                'ΤΕ01.33': 'ΤΕ02.07',
                'ΤΕ01.34': 'ΤΕ02.07',
                'ΤΕ01.35': 'ΤΕ02.07',
                'ΤΕ01.36': 'ΤΕ02.07',
                'ΔΕ01.01': 'ΔΕ02.01',
                'ΔΕ01.04': 'ΔΕ02.01',
                'ΔΕ01.07': 'ΔΕ02.01',
                'ΔΕ01.02': 'ΔΕ02.02',
                'ΔΕ01.06': 'ΔΕ02.02',
                'ΔΕ01.08': 'ΔΕ02.02',
                'ΔΕ01.10': 'ΔΕ02.02',
                'ΔΕ01.11': 'ΔΕ02.02',
                'ΔΕ01.12': 'ΔΕ02.02',
                'ΔΕ01.16': 'ΔΕ02.02',
                'ΔΕ01.13': 'ΔΕ01',
                'ΔΕ01.17': 'ΔΕ01',
                'ΔΕ02.01': 'ΔΕ02',


                '': ''
            }
            try:
                for old_string, new_string in replacements.items():
                    df['ΚΛΑΔΟΣ'] = df['ΚΛΑΔΟΣ'].str.replace(old_string, new_string, regex=True)
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method normalize_klados: {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')


def test_dieth_ekps(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    # Dictionary of key phrases and their corresponding values
    key_phrases_PE_DE = {
        'Π.Ε.',
        'Δ.Ε.'
    }

    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            # # Check and add 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column if it doesn't exist
            check_if_column_exists(df, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ', '-')
            try:
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].fillna('-', inplace=True)
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps  (replace nan to -): {e},{filename}")
            try:
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ Α/ΘΜΙΑΣ ΕΚΠ/ΣΗΣ']
                df = df.drop(['Δ/ΝΣΗ Α/ΘΜΙΑΣ ΕΚΠ/ΣΗΣ'], axis=1)
                # preprocess_Protovathmia_address_column(df)
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace Α/ΝΣΗ Α/ΘΜΙΑΣ ΕΚΠ/ΣΗΣ): {e},{filename}")
            try:
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ Π/ΘΜΙΑΣ ΕΚΠ/ΣΗΣ']
                df = df.drop(['Δ/ΝΣΗ Π/ΘΜΙΑΣ ΕΚΠ/ΣΗΣ'], axis=1)
                # preprocess_Protovathmia_address_column(df)
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps  (replace Δ/ΝΣΗ Π/ΘΜΙΑΣ ΕΚΠ/ΣΗΣ): {e},{filename}")
            try:
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΠΡΩΤΟΒΑΘΜΙΑΣ ΕΚΠ/ΣΗΣ']
                df = df.drop(['Δ/ΝΣΗ ΠΡΩΤΟΒΑΘΜΙΑΣ ΕΚΠ/ΣΗΣ'], axis=1)
                # preprocess_Protovathmia_address_column(df)
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace Δ/ΝΣΗ ΠΡΩΤΟΒΑΘΜΙΑΣ ΕΚΠ/ΣΗΣ): {e},{filename}")

            try:
                # Replace the entire value with '-' whenever 'ΠΔΕ' is found in the string
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].apply(lambda x: '-' if 'ΠΔΕ ' in x else x)
            except Exception as e:
                print(f"An error occurred during preprocessing: {e}")
            try:
                # Remove the character "ΔΙΕΥΘΥΝΣΗ" from every row in the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΙΕΥΘΥΝΣΗ', '')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace ΔΙΕΥΘΥΝΣΗ): {e},{filename}")

            try:
                df['Suffix'] = ''  # Initialize 'Suffix' before the loop
                for index, row in df.iterrows():
                    suffix_position = str(df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']).find('΄')
                    if suffix_position != -1 and suffix_position > 0:
                        # Store the character before '΄' in the 'Suffix' column
                        df.at[index, 'Suffix'] = str(df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'])[suffix_position - 1]
                        # Remove only the character before '΄'
                        df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = (
                                str(df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'])[:suffix_position - 1] +
                                str(df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'])[suffix_position:]
                        ).strip()
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace suffix): {e},{filename}")

            try:
                # Remove the character "΄" from every row in the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('΄', '')
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_dieth_ekps (replace ΄): {e},{filename}")

            # Find rows containing the word 'ΠΕΡΙΦΕΡΕΙΑΚΗ' in the column 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'

            try:
                df['Prefix'] = ''
                for index, row in df.iterrows():
                    for key_phrase in key_phrases_PE_DE:
                        if key_phrase in str(df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']):
                            # df['Prefix'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.split(' ', n=1).str[0]
                            # df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.split(' ', n=1).str[1]
                            # Store the matched phrase in the 'Prefix' column
                            matched_phrase = key_phrase

                            # Update the 'Prefix' column
                            df.at[index, 'Prefix'] = matched_phrase

                            # Remove the matched phrase from the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column
                            df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = str(df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']).replace(matched_phrase,
                                                                                                       '').strip()

                # print(df['Prefix'])
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_dieth_ekps(prefix) : {e},{filename}")

            try:
                # reconstruct the column
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Prefix'] + ' ' + df['Suffix'] + ' ' + df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']
            # df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Suffix'] + '$ ' + df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']
            except Exception as e:
                print(f"An error occurred during changing suffix: {e}")

            try:
                # Remove the character "." from every row in the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('.', '')
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_dieth_ekps (replace .): {e},{filename}")

            try:
                # Remove the character "." from every row in the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('  ', ' ')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace spaces): {e},{filename}")

            try:
                # Remove the character "." from every row in the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('()', '')
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_dieth_ekps  (replace ()): {e},{filename}")

            result = check_csv_title(filename)
            # print(filename,"<- filename, result->", result)
            global count
            count = 0
            # Iterate over each row in the DataFrame
            for index, row in df.iterrows():
                try:
                    # Convert the value to a string
                    value_str = str(row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'])
                    # Check if the value is '-' before proceeding
                    if value_str.strip() != '-':
                        # Check if 'ΔΕ' or 'ΠΕ' exists in the specific column for the current row
                        if 'ΔΕ ' not in value_str and 'ΠΕ ' not in value_str:
                            # Add 'ΔΕ' or 'ΠΕ' to the beginning of the value in the specific column for the current row
                            df.at[
                                index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = f"{'ΔΕ' if result == 'ΔΕ' else ('ΠΕ' if result == 'ΠΕ' else ('ΠΕ' if check_klados(df) else '-'))}"
                            if df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] != '-':
                                df.at[index, 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] += f" {value_str}"
                            else:
                                print(f"The file '{filename}' has '-' in row {index}")
                except Exception as e:
                    # Handle any exceptions that occur during processing
                    print(f"Error processing row {index}: {e}")
                    # Remove the character "space" from every row in the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column

            replacement_mapping = {
                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΑΤΤΙΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΤΤΙΚΗΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΑΤΤΙΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΤΤΙΚΗΣ',
                'ΠΕ ΑΤΤΙΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΤΤΙΚΗΣ',
                'ΔΕ ΑΤΤΙΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΤΤΙΚΗΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',
                "ΠΕ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ": 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',
                "ΔΕ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ": 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΕΝΤΡΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ',
                'ΠΕ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ',
                'ΔΕ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΣΤΕΡΕΑΣ ΕΛΛΑΔΑΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΚΡΗΤΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΡΗΤΗΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΚΡΗΤΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΡΗΤΗΣ',
                'ΠΕ ΚΡΗΤΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΡΗΤΗΣ',
                'ΔΕ ΚΡΗΤΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΚΡΗΤΗΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ',
                'ΠΕ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ',
                'ΔΕ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΒΟΡΕΙΟΥ ΑΙΓΑΙΟΥ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΠΕΛΟΠΟΝΝΗΣΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΠΕΛΟΠΟΝΝΗΣΟΥ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΠΕΛΟΠΟΝΝΗΣΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΠΕΛΟΠΟΝΝΗΣΟΥ',
                'ΠΕ ΠΕΛΟΠΟΝΝΗΣΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΠΕΛΟΠΟΝΝΗΣΟΥ',
                'ΔΕ ΠΕΛΟΠΟΝΝΗΣΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΠΕΛΟΠΟΝΝΗΣΟΥ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ',
                'ΠΕ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ',
                'ΔΕ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΕΛΛΑΔΑΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ',
                'ΠΕ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ',
                'ΔΕ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΑΝΑΤΟΛΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ ΚΑΙ ΘΡΑΚΗΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΗΠΕΙΡΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΗΠΕΙΡΟΥ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΗΠΕΙΡΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΗΠΕΙΡΟΥ',
                'ΠΕ ΗΠΕΙΡΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΗΠΕΙΡΟΥ',
                'ΔΕ ΗΠΕΙΡΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΗΠΕΙΡΟΥ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',
                'ΠΕ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',
                'ΔΕ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΔΥΤΙΚΗΣ ΜΑΚΕΔΟΝΙΑΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΙΟΝΙΩΝ ΝΗΣΩΝ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΙΟΝΙΩΝ ΝΗΣΩΝ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΙΟΝΙΩΝ ΝΗΣΩΝ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΙΟΝΙΩΝ ΝΗΣΩΝ',
                'ΠΕ ΙΟΝΙΩΝ ΝΗΣΩΝ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΙΟΝΙΩΝ ΝΗΣΩΝ',
                'ΔΕ ΙΟΝΙΩΝ ΝΗΣΩΝ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΙΟΝΙΩΝ ΝΗΣΩΝ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΘΕΣΣΑΛΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΘΕΣΣΑΛΙΑΣ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΘΕΣΣΑΛΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΘΕΣΣΑΛΙΑΣ',
                'ΠΕ ΘΕΣΣΑΛΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΘΕΣΣΑΛΙΑΣ',
                'ΔΕ ΘΕΣΣΑΛΙΑΣ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΘΕΣΣΑΛΙΑΣ',

                'ΔΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ',
                'ΠΕ ΠΕΡΙΦΕΡΕΙΑΚΗ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ',
                'ΠΕ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ',
                'ΔΕ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ': 'ΠΕΡΙΦΕΡΕΙΑΚΗ ΔΙΕΥΘΥΝΣΗ ΝΟΤΙΟΥ ΑΙΓΑΙΟΥ',

                'ΔΕ ΑΝ ΑΤΤΙΚΗΣ': 'ΔΕ ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ',
                'ΠΕ ΑΝ ΑΤΤΙΚΗΣ': 'ΠΕ ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ',
                'ΔΕ ΑΙΤΩΛ/ΝΙΑΣ': 'ΔΕ ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ',
                'ΠΕ ΑΙΤΩΛ/ΝΙΑΣ': 'ΠΕ ΑΙΤΩΛΟΑΚΑΡΝΑΝΙΑΣ',
                'ΔΕ ΔΥΤ ΑΤΤΙΚΗΣ': 'ΔΕ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ',

                'ΠΕ _': '-',
                'ΔΕ _': '-',

                '  ': ' ',

                'ΔΕ Α ΣΑΜΟΥ': 'ΔΕ ΣΑΜΟΥ',
                'ΠΕ Α ΣΑΜΟΥ': 'ΠΕ ΣΑΜΟΥ',
                'ΔΕ Α ΠΕΙΡΑΙΑ': 'ΔΕ ΠΕΙΡΑΙΑ',
                'ΠΕ Α ΠΕΙΡΑΙΑ': 'ΠΕ ΠΕΙΡΑΙΑ',
                'ΔΕ Α ΕΥΒΟΙΑΣ': 'ΔΕ ΕΥΒΟΙΑΣ',
                'ΠΕ Α ΕΥΒΟΙΑΣ': 'ΠΕ ΕΥΒΟΙΑΣ',
                'ΔΕ Α ΜΑΓΝΗΣΙΑΣ': 'ΔΕ ΜΑΓΝΗΣΙΑΣ',
                'ΠΕ Α ΜΑΓΝΗΣΙΑΣ': 'ΠΕ ΜΑΓΝΗΣΙΑΣ',
                'ΔΕ Β ΜΑΓΝΗΣΙΑΣ': 'ΔΕ ΜΑΓΝΗΣΙΑΣ',
                'ΠΕ Β ΜΑΓΝΗΣΙΑΣ': 'ΠΕ ΜΑΓΝΗΣΙΑΣ',
                'ΔΕ Α ΕΒΡΟΥ': 'ΔΕ ΕΒΡΟΥ',
                'ΠΕ Α ΕΒΡΟΥ': 'ΔΕ ΕΒΡΟΥ',
                'ΔΕ Α ΔΩΔΕΚΑΝΗΣΟΥ': 'ΔΕ ΔΩΔΕΚΑΝΗΣΟΥ',
                'ΠΕ Α ΔΩΔΕΚΑΝΗΣΟΥ': 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ',
                'ΠΕ Β ΔΩΔΕΚΑΝΗΣΟΥ': 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ',
                'ΔΕ Β ΔΩΔΕΚΑΝΗΣΟΥ': 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ',
                'ΠΕ Γ ΔΩΔΕΚΑΝΗΣΟΥ': 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ',
                'ΔΕ Γ ΔΩΔΕΚΑΝΗΣΟΥ': 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ',
                'ΔΕ Α ΑΝΑΤ ΑΤΤΙΚΗΣ': 'ΔΕ ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ',
                'ΠΕ Α ΑΝΑΤ ΑΤΤΙΚΗΣ': 'ΠΕ ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ',
                'ΔΕ Γ ΚΥΚΛΑΔΩΝ': 'ΔΕ ΚΥΚΛΑΔΩΝ',
                'ΠΕ Γ ΚΥΚΛΑΔΩΝ': 'ΠΕ ΚΥΚΛΑΔΩΝ',
                'ΔΕ Α ΚΕΡΚΥΡΑΣ': 'ΔΕ ΚΕΡΚΥΡΑΣ',
                'ΠΕ Α ΚΕΡΚΥΡΑΣ': 'ΠΕ ΚΕΡΚΥΡΑΣ',
                'ΔΕ Α ΚΥΚΛΑΔΩΝ': 'ΔΕ ΚΥΚΛΑΔΩΝ',
                'ΠΕ Α ΚΥΚΛΑΔΩΝ': 'ΠΕ ΚΥΚΛΑΔΩΝ',
                'ΠΕ ΔΥΤ ΑΤΤΙΚΗΣ': 'ΠΕ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ',
                'ΔΕ Α ΚΕΦΑΛΛΗΝΙΑΣ': 'ΔΕ ΚΕΦΑΛΛΗΝΙΑΣ',
                'ΠΕ Α ΚΕΦΑΛΛΗΝΙΑΣ': 'ΠΕ ΚΕΦΑΛΛΗΝΙΑΣ'
            }

            try:
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('  ', ' ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ _', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ_', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ_', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ _', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΣΑΜΟΥ', 'ΔΕ ΣΑΜΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΣΑΜΟΥ', 'ΠΕ ΣΑΜΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΠΕΙΡΑΙΑ', 'ΔΕ ΠΕΙΡΑΙΑ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΠΕΙΡΑΙΑ', 'ΠΕ ΠΕΙΡΑΙΑ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΕΥΒΟΙΑΣ', 'ΔΕ ΕΥΒΟΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΕΥΒΟΙΑΣ', 'ΠΕ ΕΥΒΟΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΜΑΓΝΗΣΙΑΣ', 'ΔΕ ΜΑΓΝΗΣΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΜΑΓΝΗΣΙΑΣ', 'ΠΕ ΜΑΓΝΗΣΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Β ΜΑΓΝΗΣΙΑΣ', 'ΔΕ ΜΑΓΝΗΣΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Β ΜΑΓΝΗΣΙΑΣ', 'ΠΕ ΜΑΓΝΗΣΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΕΒΡΟΥ', 'ΔΕ ΕΒΡΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΕΒΡΟΥ', 'ΔΕ ΕΒΡΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΔΩΔΕΚΑΝΗΣΟΥ', 'ΔΕ ΔΩΔΕΚΑΝΗΣΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΔΩΔΕΚΑΝΗΣΟΥ', 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Β ΔΩΔΕΚΑΝΗΣΟΥ', 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Β ΔΩΔΕΚΑΝΗΣΟΥ', 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Γ ΔΩΔΕΚΑΝΗΣΟΥ', 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Γ ΔΩΔΕΚΑΝΗΣΟΥ', 'ΠΕ ΔΩΔΕΚΑΝΗΣΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΑΝΑΤ ΑΤΤΙΚΗΣ', 'ΔΕ ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΑΝΑΤ ΑΤΤΙΚΗΣ', 'ΠΕ ΑΝΑΤΟΛΙΚΗΣ ΑΤΤΙΚΗΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Γ ΚΥΚΛΑΔΩΝ', 'ΔΕ ΚΥΚΛΑΔΩΝ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Γ ΚΥΚΛΑΔΩΝ', 'ΠΕ ΚΥΚΛΑΔΩΝ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΚΕΡΚΥΡΑΣ', 'ΔΕ ΚΕΡΚΥΡΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΚΕΡΚΥΡΑΣ', 'ΠΕ ΚΕΡΚΥΡΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΚΥΚΛΑΔΩΝ', 'ΔΕ ΚΥΚΛΑΔΩΝ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΚΥΚΛΑΔΩΝ', 'ΠΕ ΚΥΚΛΑΔΩΝ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ ΔΥΤ ΑΤΤΙΚΗΣ', 'ΔΕ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ ΔΥΤ ΑΤΤΙΚΗΣ', 'ΠΕ ΔΥΤΙΚΗΣ ΑΤΤΙΚΗΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΚΕΦΑΛΛΗΝΙΑΣ', 'ΔΕ ΚΕΦΑΛΛΗΝΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΚΕΦΑΛΛΗΝΙΑΣ', 'ΠΕ ΚΕΦΑΛΛΗΝΙΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΚΑΒΑΛΑΣ', 'ΔΕ ΚΑΒΑΛΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΚΑΒΑΛΑΣ', 'ΠΕ ΚΑΒΑΛΑΣ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΧΙΟΥ', 'ΔΕ ΧΙΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Α ΧΙΟΥ', 'ΠΕ ΧΙΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Β ΛΕΣΒΟΥ', 'ΔΕ ΛΕΣΒΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ Β ΛΕΣΒΟΥ', 'ΠΕ ΛΕΣΒΟΥ')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ Α ΛΕΣΒΟΥ', 'ΔΕ ΛΕΣΒΟΥ')

                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].map(replacement_mapping).fillna(df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'])

                # for old_string, new_string in replacement_mapping.items():
                #     df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace(old_string, new_string,regex=True)
                # df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ _', '-')
                # df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ _', '-')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps(replace ΠΕΡΙΦΕΡΕΙΑΚΗ names) : {e},{filename}")
            try:
                # Remove the character "space" from every row in the 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' column
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('  ', ' ')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace space 2): {e},{filename}")
            try:
                count_strings(filename, df)
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace space 2): {e},{filename}")
            try:
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df.apply(check_and_update, axis=1)
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ  -', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ -', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ-', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ  -', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ -', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ-', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΠΕ ΣΜΕΑΕ', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΔΕ ΣΜΕΑΕ', '-')
                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.replace('ΣΜΕΑΕ', '-')
            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_dieth_ekps (replace space 2): {e},{filename}")

                df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].fillna('-')

            # Drop the 'Suffix' and 'Prefix' columns
            df = df.drop(['Suffix', 'Prefix'], axis=1)

            output_path = os.path.join(folder_path, filename)

            df.to_csv(output_path, index=False, encoding='utf-8')


def count_strings(filename, df):
    # Count the number of rows containing 'Δ.Ε.' and 'Π.Ε.'
    # de_count = df.apply(lambda row: 'Δ.Ε.' in row.to_string(), axis=1).sum()
    de_count = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.contains('Δ.Ε.').sum()
    # pe_count = df.apply(lambda row: 'Π.Ε.' in row.to_string(), axis=1).sum()
    pe_count = df['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].str.contains('Π.Ε.').sum()

    # Print the results
    # print(f"The file '{filename}' has 'Δ.Ε.' {de_count} number of times")
    # print(f"The file '{filename}' has 'Π.Ε.' {pe_count} number of times")


def check_and_update(row):
    if 'ΠΕ ' in row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] or 'ΔΕ ' in row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] or 'ΠΕΡΙΦΕΡΕΙΑΚΗ' in row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']:
        return row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']
    elif 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' in row and 'Δ.Ε.' in row['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ']:
        return 'ΔΕ' + row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']
    elif 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ' in row and 'Π.Ε.' in row['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ']:
        return 'ΠΕ' + row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']
    elif 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ' in row:
        return row['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']
    else:
        return np.nan


def preprocess_Protovathmia_address_column(df):
    try:
        # Check if 'ΠΕ ' or 'ΔΕ ' exists in the values of the new column
        if not df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.contains('ΠΕ ').any() or not df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'].str.contains('ΔΕ ').any():
            # Add 'ΠΕ' to the beginning of each value
            df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ'] = 'ΠΕ ' + df['Δ/ΝΣΗ ΕΚΠ/ΣΗΣ']
    except Exception as e:
        print(f"An error occurred during preprocessing: {e}")


def check_klados(df):
    # Check if any row has the value 'ΠΕ60' or 'ΠΕ70' in the column 'ΚΛΑΔΟΣ'
    if 'ΚΛΑΔΟΣ' in df.columns:
        if 'ΠΕ60' in df['ΚΛΑΔΟΣ'].values or 'ΠΕ70' in df['ΚΛΑΔΟΣ'].values or 'ΠΕ60.ΕΑΕ' in df[
            'ΚΛΑΔΟΣ'].values or 'ΠΕ70.ΕΑΕ' in df['ΚΛΑΔΟΣ'].values:
            return True
    return False


def check_csv_title(csv_filename):
    # Define the keywords for the first and second groups
    first_group_keywords = ['ΑΘΜΙΑΣ', 'ΑΘΜΙΑ', 'Αθμιας', 'Αθμια', 'ΠΕ']
    second_group_keywords = ['ΒΘΜΙΑΣ', 'ΒΘΜΙΑ', 'Βθμιας', 'Βθμια', 'ΔΕ']

    # Convert the filename to lowercase for case-insensitive matching
    csv_filename_lower = csv_filename.lower()

    # Check for the presence of keywords in the filename
    for keyword in first_group_keywords:
        if keyword.lower() in csv_filename_lower:
            # print(csv_filename_lower,"PE case")
            return 'ΠΕ'  # Return 'ΠΕ' if first group keywords are found

    for keyword in second_group_keywords:
        if keyword.lower() in csv_filename_lower:
            # print(csv_filename_lower, "DE case")
            return 'ΔΕ'  # Return 'ΔΕ' if second group keywords are found

    # Return None if neither group keywords are found
    return None


def test_delete_periferia(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('ΠΕΡΙΦΕΡΕΙΑ', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_delete_periferia : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def test_create_sxolia(folder_path):
    # List of strings to search for
    strings_to_search = ['ΚΕΔΔΥ', 'Κ.Ε.Σ.Υ. ΚΕΔΑΣΥ 2o', 'ΚΕΔΑΣΥ 1o', 'ΚΕΔΑΣΥ', 'ΚΕΣΥ 1o', 'ΚΕΣΥ 2o', 'ΣΔΕΥ στο ΚΕΔΑΣΥ',
                         'ΣΔΕΥ ΚΕΔΑΣΥ', 'ΣΔΕΥ ΚΕΔΑΣΥ 1o', 'ΣΔΕΥ ΚΕΔΑΣΥ 2o', 'ΣΔΕΥ στο ΚΕΔΑΣΥ 2o', 'ΣΔΕΥ στο ΚΕΔΑΣΥ 1o',
                         "Κ.Ε.Σ.Υ.", "ΚΕΣΥ", "ΣΔΕΥ - ΚΕΔΑΣΥ 1o", "ΣΔΕΥ ΚΕΔΑΣΥ 1o", "ΣΔΕΥ  ", "ΣΔΕΥ  2o",
                         "Ειδικό Αγροτικό Κατάστημα Κράτησης Νέων Κασσαβέτειας [ΠΑΡΑΡΤΗΜΑ ΓΥΜΝΑΣΙΟΥ Ν. ΑΓΧΙΑΛΟΥ (Ε.Α.Κ.Κ.Ν.Κ.)]",
                         "Ειδικό Αγροτικό Σωφρονιστικό Κατάστημα Νέων Κασσαβέτειας [ΠΑΡΑΡΤΗΜΑ ΓΥΜΝΑΣΙΟΥ Ν. ΑΓΧΙΑΛΟΥ (Ε.Α.Σ.Κ.Ν.Κ.)]",
                         "Ειδικό Κατάστημα Κράτησης Νέων Αυλώνα [2ο ΓΥΜΝΑΣΙΟ ΑΥΛΩΝΑ ΜΕ ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ (Ε.Κ.Κ.Ν.Α.)]",
                         "Ειδικό Κατάστημα Κράτησης Νέων Βόλου [ΠΑΡΑΡΤΗΜΑ 6ου ΓΥΜΝΑΣΙΟΥ ΒΟΛΟΥ (Ε.Κ.Κ.Ν.Β.)]",
                         "Ειδικό Σωφρονιστικό Κατάστημα Νέων Αυλώνα [2ο ΓΥΜΝΑΣΙΟ ΑΥΛΩΝΑ ΜΕ ΛΥΚΕΙΑΚΕΣ ΤΑΞΕΙΣ (Ε.Σ.Κ.Ν.Α.)]",
                         "Ειδικό Σωφρονιστικό Κατάστημα Νέων Αυλώνα [2ο Τριθέσιο Δημοτικό Σχολείο Αυλώνας]",
                         "Ειδικό Σωφρονιστικό Κατάστημα Νέων Βόλου [ΠΑΡΑΡΤΗΜΑ 6ου ΓΥΜΝΑΣΙΟΥ ΒΟΛΟΥ (Ε.Σ.Κ.Ν.Β.)]",
                         "Σωφρονιστικό Κατάστημα Γρεβενών [Μονοθέσιο Δημοτικό Σχολείο]",
                         "ΗΡΑΚΛΕΙΟΥ – Σχολείο Ευρωπαϊκής Παιδείας", "Ξάνθη – Μειονοτικά – ελληνόγλωσσο πρόγραμμα",
                         "ΞΑΝΘΗΣ – Μειονοτικά ελληνόγλωσσο πρόγραμμα",
                         "ΞΑΝΘΗΣ – Μειονοτικά ελληνόγλωσσο πρόγραμμα", "ΡΟΔΟΠΗΣ – Μειονοτικά ελληνόγλωσσο πρόγραμμα",
                         "Α ΕΒΡΟΥ – Μειονοτικά ελληνόγλωσσο πρόγραμμα", "ΓΕΝ. ΕΚΚΛ/ΚΟ ΛΥΚΕΙΟΓΥΜΝΑΣΙΟ ΝΕΑΠΟΛΕΩΣ",
                         "Σχολείο Ευρωπαϊκής Παιδείας",
                         "ΣΔΕΥ ΚΕΔΑΣΥ 2o",
                         "Κατάστημα Κράτησης Κορίνθου [ΠΑΡΑΡΤΗΜΑ ΕΣΠΕΡΙΝΟΥ ΓΥΜΝΑΣΙΟΥ ΚΟΡΙΝΘΟΥ ΚΑΙ ΠΑΡΑΡΤΗΜΑ 1ου ΓΕΛ ΚΟΡΙΝΘΟΥ (Κ.Κ.Κ.)]"]

    def move_strings(row):
        for string in strings_to_search:
            if string in row['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ']:
                # Remove the string from 'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'
                # row['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'] = row['ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ'].replace(string, '')

                # Add the string to 'ΣΧΟΛΙΑ'
                if pd.isnull(row['ΣΧΟΛΙΑ']):
                    row['ΣΧΟΛΙΑ'] = string
                else:
                    row['ΣΧΟΛΙΑ'] = row['ΣΧΟΛΙΑ'] + ', ' + string
        return row

    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            # Check and add 'ΩΡΑΡΙΟ' column if it doesn't exist
            check_if_column_exists(df, 'ΩΡΑΡΙΟ', 'ΚΕΝΟ')

            try:
                # Create a new column 'ΣΧΟΛΙΑ' with the values of 'ΩΡΑΡΙΟ'
                df['ΣΧΟΛΙΑ'] = df['ΩΡΑΡΙΟ']
                # Drop the original 'ΩΡΑΡΙΟ' column
                df.drop(columns=['ΩΡΑΡΙΟ'], inplace=True)

                df['ΣΧΟΛΙΑ'] = df['ΣΧΟΛΙΑ'].str.replace(',', '')

                # Apply the function to each row
                df = df.apply(move_strings, axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_create_sxolia : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


# def test_create_sxolia(folder_path):
#     # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
#     for filename in os.listdir(folder_path):
#         if filename.endswith(".csv"):
#             # Construct the full path of the input CSV file
#             input_path = os.path.join(folder_path, filename)
#
#             # Read CSV file into a pandas DataFrame
#             df = pd.read_csv(input_path, encoding='utf-8')
#
#             # Check and add 'ΩΡΑΡΙΟ' column if it doesn't exist
#             check_if_column_exists(df, 'ΩΡΑΡΙΟ', 'ΚΕΝΟ')
#
#             try:
#                 # Create a new column 'ΣΧΟΛΙΑ' with the values of 'ΩΡΑΡΙΟ'
#                 df['ΣΧΟΛΙΑ'] = df['ΩΡΑΡΙΟ']
#                 # Drop the original 'ΩΡΑΡΙΟ' column
#                 df.drop(columns=['ΩΡΑΡΙΟ'], inplace=True)
#             except Exception as e:
#                 print(f"An error occurred during preprocessing on method test_create_sxolia : {e},{filename}")
#
#             # Construct the full path of the output CSV file
#             output_path = os.path.join(folder_path, filename)
#
#             # Save the modified DataFrame back to a new CSV file
#             df.to_csv(output_path, index=False, encoding='utf-8')
#
#             # Optionally, you can delete the original CSV file
#             # os.remove(input_path)


def test_add_mousika_organa_to_sxolia(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Create a new column 'ΣΧΟΛΙΑ' by concatenating values from 'ΣΧΟΛΙΑ' and 'ΜΟΥΣΙΚΗ ΕΙΔΙΚΕΥΣΗ' with a comma
                df['ΣΧΟΛΙΑ'] = df['ΣΧΟΛΙΑ'] + ', ' + df['ΜΟΥΣΙΚΗ ΕΙΔΙΚΕΥΣΗ']  # prosthetei "" stin teleytaia stili
                # Drop the original columns 'ΩΡΑΡΙΟ' and 'ΜΟΥΣΙΚΗ ΕΙΔΙΚΕΥΣΗ'
                df.drop(columns=['ΜΟΥΣΙΚΗ ΕΙΔΙΚΕΥΣΗ'], inplace=True)
                df.drop(columns=['ΩΡΑΡΙΟ'], inplace=True)

            except Exception as e:
                print(
                    f"An error occurred during preprocessing on method test_add_mousika_organa_to_sxolia : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')

            # Optionally, you can delete the original CSV file
            # os.remove(input_path)


def test_add_hmeromhnia(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli\preprocess_folder"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            try:
                # Extract the date from the filename
                match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
                date_from_filename = match.group() if match else None

                if date_from_filename:
                    # Construct the full path of the input CSV file
                    input_path = os.path.join(folder_path, filename)

                    # Read CSV file into a pandas DataFrame
                    df = pd.read_csv(input_path, encoding='utf-8')

                    # Create a new column and assign the corresponding date
                    df['ΗΜΕΡΟΜΗΝΙΑ'] = date_from_filename
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_add_hmeromhnia : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')


def test_create_sxoliko_etos(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli\preprocess_folder"
    global date_object
    # date_object = datetime.today()
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):

            # Extract the date from the filename
            match = re.search(r'\d{4}-\d{2}-\d{2}', filename)
            date_from_filename = match.group() if match else None

            # Extract the year from the filename
            match = re.search(r'\d{4}', filename)
            year_from_filename = match.group() if match else None
            try:
                date_object = datetime.strptime(date_from_filename, "%Y-%m-%d").date()
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_create_sxoliko_etos : {e},{filename}")

            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')
            if year_from_filename:

                try:
                    # Check if the date is in the range from September 1st to December 31st
                    if 1 <= date_object.month <= 6:
                        # Create the 'ΣΧΟΛΙΚΟ ΕΤΟΣ' column
                        df['ΕΤΟΣ'] = f"{int(year_from_filename) - 1}-{year_from_filename}"
                        # print("Date is in the range January 1st - August 31st")
                    elif 7 <= date_object.month <= 12:
                        # Create the 'ΣΧΟΛΙΚΟ ΕΤΟΣ' column
                        df['ΕΤΟΣ'] = f"{year_from_filename}-{int(year_from_filename) + 1}"

                        # print("Date is in the range September 1st - December 31st")
                    else:
                        print("Date is not in either range")
                except Exception as e:
                    print(
                        f"An error occurred during preprocessing on method test_create_sxoliko_etos : {e},{filename}, or date {date_object}")

                # Construct the full path of the output CSV file
                output_path = os.path.join(folder_path, filename)

                # Save the modified DataFrame back to a new CSV file
                df.to_csv(output_path, index=False, encoding='utf-8')


def test_add_orario_values(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli\preprocess_folder"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv") and "ΑΠΩ" in filename:
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')

            if "ΩΡΑΡΙΟ" not in df.columns:
                # Add the 'ΩΡΑΡΙΟ' column with 'ΑΠΩ' as the default value
                df['ΩΡΑΡΙΟ'] = 'ΑΠΩ'

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')


def normalize_klados(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli\preprocess_folder"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            df = pd.read_csv(input_path, encoding='utf-8')
            # Check and add 'ΤΥΠΟΣ' column if it doesn't exist
            check_if_column_exists(df, 'ΚΛΑΔΟΣ', '')
            try:
                # Check if the 'ΚΛΑΔΟΣ/ ΕΙΔΙΚΟΤΗΤΑ' column exists in the DataFrame
                if 'ΚΛΑΔΟΣ/ ΕΙΔΙΚΟΤΗΤΑ' in df.columns:
                    # Keep only the 'ΚΛΑΔΟΣ/ ΕΙΔΙΚΟΤΗΤΑ' column
                    df['ΚΛΑΔΟΣ'] = df['ΚΛΑΔΟΣ/ ΕΙΔΙΚΟΤΗΤΑ']
                    df = df.drop(['ΚΛΑΔΟΣ/ ΕΙΔΙΚΟΤΗΤΑ'], axis=1)
                elif 'ΚΛΑΔΟΣ / ΕΙΔΙΚΟΤΗΤΑ' in df.columns:
                    # Keep only the 'ΚΛΑΔΟΣ/ ΕΙΔΙΚΟΤΗΤΑ' column
                    df['ΚΛΑΔΟΣ'] = df['ΚΛΑΔΟΣ / ΕΙΔΙΚΟΤΗΤΑ']
                    df = df.drop(['ΚΛΑΔΟΣ / ΕΙΔΙΚΟΤΗΤΑ'], axis=1)
                elif 'ΕΙΔΙΚΟΤΗΤΑ' in df.columns:
                    df['ΚΛΑΔΟΣ'] = df['ΕΙΔΙΚΟΤΗΤΑ']
                    df = df.drop(['ΕΙΔΙΚΟΤΗΤΑ'], axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_normalize_columns : {e},{filename}")

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')


def test_normalize_columns(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli\preprocess_folder"
    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)

            df = pd.read_csv(input_path, encoding='utf-8')

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('ΣΧΟΛΙΚΗ ΜΟΝΑΔΑ ΤΟΠΟΘΕΤΗΣΗΣ', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_normalize_columns : {e},{filename}")

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('ΣΧΟΛΙΚΗ ΜΟΝΑΔΑ ΤΟΠΟΘΕΤΗΣΗΣ', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_normalize_columns : {e},{filename}")

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('Τύπος Πρόσληψης', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_normalize_columns : {e},{filename}")

            try:
                # Remove the column named 'ΜΗΤΡΩΝΥΜΟ'
                df = df.drop('ΤΡΙΤΕΚΝΟΣ', axis=1)
            except Exception as e:
                print(f"An error occurred during preprocessing on method test_normalize_columns : {e},{filename}")
            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')


def test_normalize_all_columns_names(folder_path):
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli\preprocess_folder"

    for filename in os.listdir(folder_path):
        if filename.endswith(".csv"):
            # Construct the full path of the input CSV file
            input_path = os.path.join(folder_path, filename)
            column_mapping = {
                'ΤΥΠΟΣ': 'Typos',
                'ΕΠΩΝΥΜΟ': 'Eponymo',
                'ΟΝΟΜΑ': 'Onoma',
                'ΠΑΤΡΩΝΥΜΟ': 'Patronymo',
                'ΚΛΑΔΟΣ': 'Klados',
                'ΣΕΙΡΑ ΠΙΝΑΚΑ': 'Seira_Pinaka',
                'ΜΟΡΙΑ ΠΙΝΑΚΑ': 'Moria_Pinaka',
                'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ': 'Perioxh_Topothethshs',
                'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ': 'Dieytynsh_Ekpaideyshs',
                'ΗΜΕΡΟΜΗΝΙΑ': 'Hmeromnia',
                'ΕΤΟΣ': 'Etos',
                'ΣΧΟΛΙΑ': 'Sxolia',
            }
            # Read CSV file into a pandas DataFrame
            df = pd.read_csv(input_path, encoding='utf-8')
            try:
                # Keep only the columns present in the mapping dictionary
                df = df[list(column_mapping.keys())]

                # Rename the columns to English names
                df.columns = [column_mapping[col] for col in df.columns]

            except Exception as e:

                print(f"An error occurred during preprocessing on method test_normalize_all_columns : {e},{filename}")
                # Check if all desired columns exist in the DataFrame

            # Construct the full path of the output CSV file
            output_path = os.path.join(folder_path, filename)

            # Save the modified DataFrame back to a new CSV file
            df.to_csv(output_path, index=False, encoding='utf-8')


def test_full_outer_join_csv_files(folder_path, final_output_path):
    global output_path
    # folder_path = r"C:\Users\mike2\OneDrive\Desktop\sxoli"
    # Initialize an empty DataFrame to store the result
    result_df = None
    try:
        # Get the current date
        current_date = date.today()

        # Generate a random 5-digit number
        random_number = random.randint(10000, 99999)

        for filename in os.listdir(folder_path):
            if filename.endswith(".csv"):
                # Construct the full path of the input CSV file
                input_path = os.path.join(folder_path, filename)

                # Read CSV file into a pandas DataFrame
                current_df = pd.read_csv(input_path, encoding='utf-8')

                # Fill all NaN values in the current DataFrame with ''
                # current_df = current_df.fillna('')
                # Select string columns and fill NaN values with ''
                string_columns = current_df.select_dtypes(include=['object']).columns
                current_df[string_columns] = current_df[string_columns].fillna('')

                current_df['Moria_Pinaka'] = pd.to_numeric(current_df['Moria_Pinaka'], errors='coerce')
                current_df['Moria_Pinaka'] = current_df['Moria_Pinaka'].astype('float64')

                current_df['Moria_Pinaka'] = current_df['Moria_Pinaka'].replace('', 'np.nan')

                current_df['Perioxh_Topothethshs'] = current_df['Perioxh_Topothethshs'].astype('object')
                current_df['Klados'] = current_df['Klados'].astype('object')

                current_df['Seira_Pinaka'] = pd.to_numeric(current_df['Seira_Pinaka'], errors='coerce').fillna(
                    0).astype(int)
                # Perform a full outer join with the result DataFrame
                if result_df is None:
                    result_df = current_df
                else:
                    try:
                        # print(result_df.dtypes)
                        # print(current_df.dtypes)
                        result_df = result_df.fillna(np.nan)
                        result_df = pd.merge(result_df, current_df, how='outer')
                    except Exception as e:
                        print(f"An error occurred during preprocessing on method test_full_outer_join_csv_files : {e}")
                    except pd.errors.MergeError as merge_error:
                        # Print a warning message and continue with the next file
                        warnings.warn(f"Warning in file {filename}: {merge_error}")
                # Construct the full path of the output CSV file with the desired format
                output_filename = f"merged_output-{current_date}-{random_number}.csv"
                output_path = os.path.join(final_output_path, output_filename)

        # Remove duplicate rows
        result_df = result_df.drop_duplicates()

        # Save the result DataFrame to a new CSV file
        result_df.to_csv(output_path, index=False, encoding='utf-8', na_rep='NULL')

    except Exception as e:
        print(f"An error occurred during preprocessing on method test_full_outer_join_csv_files : {e}")


def re_order(folder_path):
    desired_order = ['ΤΥΠΟΣ', 'ΕΠΩΝΥΜΟ', 'ΟΝΟΜΑ', 'ΠΑΤΡΩΝΥΜΟ', 'ΚΛΑΔΟΣ', 'ΣΕΙΡΑ ΠΙΝΑΚΑ',
                     'ΜΟΡΙΑ ΠΙΝΑΚΑ',
                     'ΠΕΡΙΟΧΗ ΤΟΠΟΘΕΤΗΣΗΣ', 'Δ/ΝΣΗ ΕΚΠ/ΣΗΣ', 'ΗΜΕΡΟΜΗΝΙΑ', 'ΕΤΟΣ', 'ΣΧΟΛΙΑ']
    try:
        for filename in os.listdir(folder_path):
            if filename.endswith(".csv"):
                # Construct the full path of the input CSV file
                input_path = os.path.join(folder_path, filename)

                # Read CSV file into a pandas DataFrame
                df = pd.read_csv(input_path, encoding='utf-8')

                # Handle missing columns
                missing_columns = set(desired_order) - set(df.columns)
                for missing_column in missing_columns:
                    df[missing_column] = None  # You can modify this to assign a default value

                # Remove extra columns
                extra_columns = set(df.columns) - set(desired_order)
                df = df.drop(columns=extra_columns, errors='ignore')

                # Reorder columns based on the desired order
                df = df[desired_order]

                # Reorder columns based on the desired order
                df = df[desired_order]

                # Construct the full path of the output CSV file
                output_path = os.path.join(folder_path, filename)

                # Save the modified DataFrame back to a new CSV file
                df.to_csv(output_path, index=False, encoding='utf-8')
    except Exception as e:
        print(f"An error occurred during preprocessing on method test_full_outer_join_csv_files : {e}")


def remove_rows_with_empty_names(folder_path):
    try:
        for filename in os.listdir(folder_path):
            if filename.endswith(".csv"):
                # Construct the full path of the input CSV file
                input_path = os.path.join(folder_path, filename)

                # Read CSV file into a pandas DataFrame
                df = pd.read_csv(input_path, encoding='utf-8')

                # Check if 'ΕΠΩΝΥΜΟ', 'ΟΝΟΜΑ', and 'ΠΑΤΡΩΝΥΜΟ' columns are empty
                condition = df['Eponymo'].isnull() & df['Onoma'].isnull() & df['Patronymo'].isnull()

                # Remove rows that meet the condition
                df = df[~condition]

                # Save the modified DataFrame back to the original CSV file (overwrite)
                df.to_csv(input_path, index=False, encoding='utf-8')

    except Exception as e:
        print(f"An error occurred during preprocessing on method remove_rows_with_empty_names : {e}")
